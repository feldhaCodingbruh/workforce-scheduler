
import calendar
import csv
from datetime import date, datetime
from difflib import SequenceMatcher
from io import BytesIO, StringIO
import json
import os
import re
import unicodedata
import urllib.error
import urllib.request
from copy import deepcopy
from pathlib import Path
from uuid import uuid4

from flask import Flask, abort, redirect, render_template, request, send_file, url_for
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

DATA_FILE = Path(__file__).with_name("scheduler_data.json")
EXPORTS_DIR = Path(__file__).with_name("exports")
DEFAULT_MONTH = 3
DEFAULT_YEAR = 2026
DEFAULT_FULL_TIME_HOURS = 160
RULES_VERSION = 6
BUILD_NUMBER = "0.7.1"
DEFAULT_AI_MODEL = os.getenv("OPENAI_MODEL", "gpt-5-mini")
FULL_SHIFT_MIN_HOURS = 7.0
HALF_SHIFT_HOURS = 4.0
MIN_SHORT_SHIFT_MINUTES = 270
MIN_REST_HOURS = 11
MAX_CONSECUTIVE_DAYS = 5
# Conservative soft preferences calibrated from anonymized schedule history.
PREFERRED_REST_HOURS = 13
SOFT_TURNAROUND_PENALTY_PER_HOUR = 1.0
ADJACENT_START_TOLERANCE_MINUTES = 60
ADJACENT_START_BONUS = 0.15
MAX_IMPORT_BYTES = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 2000
MAX_IMPORT_COLUMNS = 20
MAX_IMPORT_SHIFT_COLUMNS = 5

LOCATION_CONFIGS = [
    {"id": "location-a", "name": "Location A"},
    {"id": "location-b", "name": "Location B"},
    {"id": "location-c", "name": "Location C"},
    {"id": "location-d", "name": "Location D"},
    {"id": "location-e", "name": "Location E"},
    {"id": "location-f", "name": "Location F"},
    {"id": "location-g", "name": "Location G"},
    {"id": "location-h", "name": "Location H"},
    {"id": "location-i", "name": "Location I"},
]
DEFAULT_LOCATION_ID = LOCATION_CONFIGS[0]["id"]
DAY_SHORT_NAMES = ["Pr", "An", "Tr", "Kt", "Pn", "St", "Sk"]
DAY_LONG_NAMES = [
    "Pirmadienis", "Antradienis", "Treciadienis", "Ketvirtadienis",
    "Penktadienis", "Sestadienis", "Sekmadienis",
]
MONTH_NAMES = {
    1: "Sausis", 2: "Vasaris", 3: "Kovas", 4: "Balandis", 5: "Geguze", 6: "Birzelis",
    7: "Liepa", 8: "Rugpjutis", 9: "Rugsejis", 10: "Spalis", 11: "Lapkritis", 12: "Gruodis",
}


def build_shift(start, end, hours, label):
    return {"label": label, "start": start, "end": end, "hours": float(hours)}


def build_day_rules(shift_specs, breaks=None):
    return {
        "shifts": [build_shift(start, end, hours, f"Pamaina {index}") for index, (start, end, hours) in enumerate(shift_specs, start=1)],
        "breaks": list(breaks or []),
    }


DEFAULT_LOCATION_RULE = {
    "name": "Bazinis sablonas",
    "notes": ["Hard-coded taisykles.", "Grafikas remiasi dienos poreikiu."],
    "days": [
        build_day_rules([("10:00", "18:30", 8), ("13:00", "21:30", 8)], ["13:30-14:00", "16:30-17:00"]),
        build_day_rules([("10:00", "18:30", 8), ("13:00", "21:30", 8)], ["13:30-14:00", "16:30-17:00"]),
        build_day_rules([("10:00", "18:30", 8), ("13:00", "21:30", 8)], ["13:30-14:00", "16:30-17:00"]),
        build_day_rules([("10:00", "18:30", 8), ("13:00", "21:30", 8)], ["13:30-14:00", "16:30-17:00"]),
        build_day_rules([("10:00", "18:30", 8), ("13:00", "21:30", 8), ("15:00", "21:30", 6)], ["13:30-14:00", "16:30-17:00", "17:00-17:30"]),
        build_day_rules([("10:00", "18:30", 8), ("13:00", "21:30", 8), ("15:00", "21:30", 6)], ["13:30-14:00", "16:30-17:00", "17:00-17:30"]),
        build_day_rules([("10:00", "18:30", 8), ("13:00", "21:30", 8), ("15:00", "21:30", 6)], ["13:30-14:00", "16:30-17:00", "17:00-17:30"]),
    ],
}

LOCATION_RULES = {
    "location-a": {"name": "Location A", "notes": ["Sample three-shift template"], "days": [
        build_day_rules([("09:00", "17:30", 8), ("13:00", "21:30", 8), ("13:00", "21:30", 8)], ["13:00-13:30", "16:30-17:00", "17:00-17:30"]),
        build_day_rules([("09:00", "17:30", 8), ("13:00", "21:30", 8), ("13:00", "21:30", 8)], ["13:00-13:30", "16:30-17:00", "17:00-17:30"]),
        build_day_rules([("09:00", "17:30", 8), ("13:00", "21:30", 8), ("13:00", "21:30", 8)], ["13:00-13:30", "16:30-17:00", "17:00-17:30"]),
        build_day_rules([("09:00", "17:30", 8), ("13:00", "21:30", 8), ("13:00", "21:30", 8)], ["13:00-13:30", "16:30-17:00", "17:00-17:30"]),
        build_day_rules([("09:00", "17:30", 8), ("13:00", "21:30", 8), ("13:00", "21:30", 8), ("15:00", "21:30", 6)], ["13:00-13:30", "16:30-17:00", "17:00-17:30"]),
        build_day_rules([("09:00", "17:30", 8), ("12:00", "20:30", 8), ("13:00", "21:30", 8), ("13:00", "21:30", 8)], ["13:00-13:30", "16:00-16:30", "16:30-17:00", "17:00-17:30"]),
        build_day_rules([("09:00", "17:30", 8), ("12:00", "20:30", 8), ("13:00", "21:30", 8)], ["13:00-13:30", "16:00-16:30", "16:30-17:00"]),
    ]},
    "location-b": {"name": "Location B", "notes": ["Sample mixed-hours template"], "days": [
        build_day_rules([("10:00", "18:30", 8), ("13:00", "21:30", 8)], ["13:00-13:30", "17:00-17:30"]),
        build_day_rules([("10:00", "18:30", 8), ("13:00", "21:30", 8)], ["13:00-13:30", "17:00-17:30"]),
        build_day_rules([("10:00", "18:30", 8), ("13:00", "21:30", 8)], ["13:00-13:30", "17:00-17:30"]),
        build_day_rules([("10:00", "17:30", 7), ("14:00", "22:30", 8)], ["13:00-13:30", "16:30-17:00"]),
        build_day_rules([("10:00", "18:30", 8), ("14:00", "22:30", 8), ("18:30", "22:30", 4)], ["14:00-14:30", "17:30-18:00", "18:00-18:30"]),
        build_day_rules([("10:00", "18:30", 8), ("14:00", "22:30", 8), ("14:00", "22:30", 8)], ["14:00-14:30", "17:30-18:00", "18:00-18:30"]),
        build_day_rules([("10:00", "18:30", 8), ("13:00", "21:30", 8), ("13:00", "21:30", 8)], ["13:00-13:30", "16:30-17:00", "17:30-18:00"]),
    ]},
    "location-c": {"name": "Location C", "notes": ["Sample two-shift template"], "days": [
        build_day_rules([("10:00", "17:30", 7), ("13:00", "21:30", 8), ("14:00", "21:30", 7)], ["13:30-14:00", "16:30-17:00", "17:00-17:30"]),
        build_day_rules([("10:00", "17:30", 7), ("13:00", "21:30", 8)], ["13:30-14:00", "16:30-17:00"]),
        build_day_rules([("10:00", "17:30", 7), ("13:00", "21:30", 8)], ["13:30-14:00", "16:30-17:00"]),
        build_day_rules([("10:00", "17:30", 7), ("13:00", "21:30", 8)], ["13:30-14:00", "16:30-17:00"]),
        build_day_rules([("10:00", "17:30", 7), ("13:00", "21:30", 8), ("13:00", "21:30", 8)], ["13:30-14:00", "16:30-17:00", "17:00-17:30"]),
        build_day_rules([("10:00", "17:30", 7), ("13:00", "21:30", 8)], ["13:30-14:00", "16:30-17:00"]),
        build_day_rules([("10:00", "17:30", 7), ("13:00", "21:30", 8)], ["13:30-14:00", "16:30-17:00"]),
    ]},
    "location-d": {"name": "Location D", "notes": ["Sample base template"], "days": DEFAULT_LOCATION_RULE["days"]},
    "location-e": {"name": "Location E", "notes": ["Sample base template"], "days": DEFAULT_LOCATION_RULE["days"]},
    "location-f": {"name": "Location F", "notes": ["Sample base template"], "days": DEFAULT_LOCATION_RULE["days"]},
    "location-g": DEFAULT_LOCATION_RULE,
    "location-h": DEFAULT_LOCATION_RULE,
    "location-i": DEFAULT_LOCATION_RULE,
}


def normalize_text(text):
    normalized = unicodedata.normalize("NFKD", str(text or "").strip().lower())
    return "".join(character for character in normalized if not unicodedata.combining(character))


def safe_float(value, default=0.0):
    try:
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return default


def sanitize_int(value, default, minimum, maximum):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    return parsed if minimum <= parsed <= maximum else default


def is_valid_time(time_str):
    if not isinstance(time_str, str):
        return False
    match = re.fullmatch(r"(\d{2}):(\d{2})", time_str.strip())
    if not match:
        return False
    hour = int(match.group(1))
    minute = int(match.group(2))
    return 0 <= hour <= 23 and 0 <= minute <= 59


TIME_RANGE_PATTERN = r"(?<![\d:.])(\d{1,2})(?:[:.](\d{2}))?\s*[-\u2013\u2014]\s*(\d{1,2})(?:[:.](\d{2}))?(?![\d:.])"
TIME_RANGE_SEARCH_RE = re.compile(TIME_RANGE_PATTERN)
TIME_RANGE_FULL_RE = re.compile(rf"\s*{TIME_RANGE_PATTERN}\s*")


def interval_from_match(match):
    start_hour, start_minute, end_hour, end_minute = match.groups()
    start_time = f"{int(start_hour):02d}:{int(start_minute or 0):02d}"
    end_time = f"{int(end_hour):02d}:{int(end_minute or 0):02d}"
    if not is_valid_time(start_time) or not is_valid_time(end_time):
        return None
    if time_to_minutes(start_time) >= time_to_minutes(end_time):
        return None
    return start_time, end_time


def time_to_minutes(time_str):
    hour, minute = map(int, time_str.split(":"))
    return hour * 60 + minute


def minutes_to_time(minutes):
    return f"{minutes // 60:02d}:{minutes % 60:02d}"


def shift_length_hours(start, end):
    return (time_to_minutes(end) - time_to_minutes(start)) / 60


def etatas_to_month_hours(etatas, full_time_hours=DEFAULT_FULL_TIME_HOURS):
    return safe_float(etatas) * full_time_hours


def get_days_in_month(year, month):
    return calendar.monthrange(year, month)[1]


def get_day_type(year, month, day):
    weekday_index = calendar.weekday(year, month, day)
    return ("weekend", weekday_index) if weekday_index >= 5 else ("weekday", weekday_index)


def weekday_name_from_index(index):
    return DAY_SHORT_NAMES[index]


def format_demand_value(value):
    rounded = round(value * 2) / 2
    return str(int(round(rounded))) if abs(rounded - round(rounded)) < 0.001 else str(rounded).replace(".", ",")


def sanitize_demand_value(value, default):
    parsed = safe_float(value, default)
    return round(max(0.0, min(parsed, 5.0)) * 2) / 2


def get_location_rule(location_id):
    return LOCATION_RULES.get(location_id, DEFAULT_LOCATION_RULE)


def get_day_rules_for_weekday(location_id, weekday_index):
    return deepcopy(get_location_rule(location_id)["days"][weekday_index])


def split_template_shifts(day_rules):
    full_shifts, half_shifts = [], []
    for shift in day_rules["shifts"]:
        copied = deepcopy(shift)
        if copied["hours"] >= FULL_SHIFT_MIN_HOURS:
            full_shifts.append(copied)
        else:
            half_shifts.append(copied)
    if not full_shifts and day_rules["shifts"]:
        full_shifts = [deepcopy(item) for item in sorted(day_rules["shifts"], key=lambda item: (-item["hours"], item["start"]))]
    if not half_shifts and full_shifts:
        source_shift = max(full_shifts, key=lambda item: time_to_minutes(item["end"]))
        end_minutes = time_to_minutes(source_shift["end"])
        start_minutes = max(time_to_minutes(source_shift["start"]), end_minutes - int(HALF_SHIFT_HOURS * 60))
        half_shifts.append({"label": "Puse", "start": minutes_to_time(start_minutes), "end": source_shift["end"], "hours": shift_length_hours(minutes_to_time(start_minutes), source_shift["end"])})
    return full_shifts, half_shifts


def build_connected_half_shift(day_rules):
    all_shifts = day_rules["shifts"]
    if not all_shifts:
        return None
    opening_start = min(time_to_minutes(shift["start"]) for shift in all_shifts)
    closing_end = max(time_to_minutes(shift["end"]) for shift in all_shifts)
    half_start = max(opening_start, closing_end - int(HALF_SHIFT_HOURS * 60))
    return {
        "label": "Puse",
        "start": minutes_to_time(half_start),
        "end": minutes_to_time(closing_end),
        "hours": shift_length_hours(minutes_to_time(half_start), minutes_to_time(closing_end)),
    }


def build_standard_opening_full_shift(day_rules, year, month, day):
    all_shifts = day_rules["shifts"]
    if not all_shifts:
        return None
    opening_start = min(time_to_minutes(shift["start"]) for shift in all_shifts)
    weekday_index = calendar.weekday(year, month, day)
    default_end = "18:00" if weekday_index == 6 else "18:30"
    target_end = time_to_minutes(default_end)
    return {
        "label": "Pilna 1",
        "start": minutes_to_time(opening_start),
        "end": default_end,
        "hours": shift_length_hours(minutes_to_time(opening_start), default_end),
    }


def adjust_shift_for_special_opening(year, month, day, shift, index):
    weekday_index = calendar.weekday(year, month, day)
    if weekday_index != 6 or index != 0:
        return shift

    adjusted = deepcopy(shift)
    adjusted["start"] = "09:30"
    if adjusted["role"] == "full":
        adjusted["end"] = "18:00"
        adjusted["hours"] = 8.0
    return adjusted


def get_full_shift_for_requested_index(full_templates, index):
    if index < len(full_templates):
        return deepcopy(full_templates[index]), False
    if not full_templates:
        return None, False

    source_shift = max(
        full_templates,
        key=lambda item: (time_to_minutes(item["end"]), time_to_minutes(item["start"])),
    )
    return deepcopy(source_shift), True


def build_default_demand_values(location_id, year, month):
    values = []
    for day in range(1, get_days_in_month(year, month) + 1):
        weekday_index = calendar.weekday(year, month, day)
        full_shifts, half_shifts = split_template_shifts(get_day_rules_for_weekday(location_id, weekday_index))
        values.append(len(full_shifts) + (0.5 if half_shifts else 0.0))
    return values


def build_default_demand_raw(location_id, year, month):
    return "\n".join(format_demand_value(value) for value in build_default_demand_values(location_id, year, month))


def sanitize_schedule_settings(raw_settings):
    raw_settings = raw_settings if isinstance(raw_settings, dict) else {}
    return {
        "month": sanitize_int(raw_settings.get("month"), DEFAULT_MONTH, 1, 12),
        "year": sanitize_int(raw_settings.get("year"), DEFAULT_YEAR, 2000, 2100),
        "full_time_hours": sanitize_int(
            raw_settings.get("full_time_hours"),
            DEFAULT_FULL_TIME_HOURS,
            1,
            250,
        ),
    }


def normalize_availability_raw(raw_text):
    return str(raw_text or "").replace("\r\n", "\n").replace("\r", "\n")


def sanitize_workers(raw_workers):
    workers = []
    for worker in raw_workers if isinstance(raw_workers, list) else []:
        if not isinstance(worker, dict):
            continue
        name = str(worker.get("name") or "").strip()
        etatas = str(worker.get("etatas") or "").strip()
        if not name or not etatas:
            continue
        workers.append({"id": str(worker.get("id") or uuid4().hex), "name": name, "etatas": etatas, "availability_raw": str(worker.get("availability_raw") or "")})
    return workers


def sanitize_schedule_insights(raw_insights):
    if not isinstance(raw_insights, dict) or not isinstance(raw_insights.get("items"), list):
        return None
    return {"items": [str(item) for item in raw_insights["items"] if str(item).strip()], "ai_used": bool(raw_insights.get("ai_used")), "ai_status": str(raw_insights.get("ai_status") or "").strip()}


def sanitize_demand_raw(raw_text, location_id, schedule_settings):
    year, month = schedule_settings["year"], schedule_settings["month"]
    defaults = build_default_demand_values(location_id, year, month)
    lines = [line.strip() for line in str(raw_text or "").splitlines()]
    normalized = []
    for index in range(get_days_in_month(year, month)):
        raw_value = lines[index] if index < len(lines) else defaults[index]
        normalized.append(format_demand_value(sanitize_demand_value(raw_value, defaults[index])))
    return "\n".join(normalized)


def sanitize_location(raw_location, config):
    raw_location = raw_location if isinstance(raw_location, dict) else {}
    schedule_settings = sanitize_schedule_settings(raw_location.get("schedule_settings"))
    generated_schedule = raw_location.get("generated_schedule") if isinstance(raw_location.get("generated_schedule"), list) else []
    return {
        "name": str(raw_location.get("name") or config["name"]),
        "schedule_settings": schedule_settings,
        "demand_raw": sanitize_demand_raw(raw_location.get("demand_raw", ""), config["id"], schedule_settings),
        "workers": sanitize_workers(raw_location.get("workers")),
        "generated_schedule": generated_schedule,
        "worker_summary": raw_location.get("worker_summary") if isinstance(raw_location.get("worker_summary"), list) else None,
        "schedule_insights": sanitize_schedule_insights(raw_location.get("schedule_insights")),
    }


def load_app_data():
    raw_data = {}
    if DATA_FILE.exists():
        try:
            raw_data = json.loads(DATA_FILE.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            raw_data = {}
    raw_locations = raw_data.get("locations", {}) if isinstance(raw_data, dict) else {}
    stored_rules_version = raw_data.get("rules_version") if isinstance(raw_data, dict) else None
    locations = {config["id"]: sanitize_location(raw_locations.get(config["id"]), config) for config in LOCATION_CONFIGS}
    if stored_rules_version != RULES_VERSION:
        for location in locations.values():
            location["generated_schedule"] = []
            location["worker_summary"] = None
            location["schedule_insights"] = None
    return {"rules_version": RULES_VERSION, "locations": locations}


def save_app_data():
    DATA_FILE.write_text(json.dumps(app_data, ensure_ascii=False, indent=2), encoding="utf-8")


app_data = load_app_data()
save_app_data()
EXPORTS_DIR.mkdir(exist_ok=True)


def parse_availability_lines(raw_text):
    normalized = normalize_availability_raw(raw_text)
    if not normalized:
        return []
    return [line.strip() for line in normalized.split("\n")]


def parse_daily_demand_map(raw_text, year, month, location_id):
    defaults = build_default_demand_values(location_id, year, month)
    lines = [line.strip() for line in str(raw_text or "").splitlines()]
    demand_map, demand_rows = {}, []
    for day in range(1, get_days_in_month(year, month) + 1):
        raw_value = lines[day - 1] if day - 1 < len(lines) else defaults[day - 1]
        value = sanitize_demand_value(raw_value, defaults[day - 1])
        demand_map[day] = value
        demand_rows.append({"day": day, "weekday_name": weekday_name_from_index(calendar.weekday(year, month, day)), "value": value, "label": format_demand_value(value), "requested_hours": int(value * 8)})
    return demand_map, demand_rows


def get_worker_status(day_count, expected_days, filled_day_count=None):
    filled_day_count = day_count if filled_day_count is None else filled_day_count
    if day_count > expected_days:
        return f"Per daug: +{day_count - expected_days} d."
    if filled_day_count < expected_days:
        return f"Truksta {expected_days - filled_day_count} d."
    if day_count == expected_days:
        return "Gerai"
    return f"Truksta {expected_days - day_count} d."


def extract_time(text):
    match = re.search(r"(\d{1,2})(?:[:.](\d{2}))?", text)
    if not match:
        return None
    hour = int(match.group(1))
    minute = int(match.group(2)) if match.group(2) else 0
    return f"{hour:02d}:{minute:02d}" if 0 <= hour <= 23 and 0 <= minute <= 59 else None


def extract_all_times(text):
    times = []
    for hour_text, minute_text in re.findall(r"(\d{1,2})(?:[:.](\d{2}))?", text):
        hour = int(hour_text)
        minute = int(minute_text) if minute_text else 0
        if 0 <= hour <= 23 and 0 <= minute <= 59:
            times.append(f"{hour:02d}:{minute:02d}")
    return times


def build_unknown_parse_result(original, parsed_text="Neatpazinta"):
    return {"original": original.strip(), "type": "unknown", "from_time": None, "until_time": None, "second_from_time": None, "preference": None, "parsed_text": parsed_text, "parser_source": "rules", "confidence": None, "ai_notes": None}


def parse_single_line_rule_based(line):
    original = line.strip()
    normalized = normalize_text(original)
    normalized_clean = re.sub(r"\s+", " ", normalized).strip(" .,-")
    result = build_unknown_parse_result(original)
    if any(word in normalized for word in ["rytas", "rytine", "rytines", "geriau rytas", "noriu rytines"]):
        result["preference"] = "morning"
    elif any(word in normalized for word in ["vakaras", "vakarine", "vakaro", "geriau vakaras"]):
        result["preference"] = "evening"
    if normalized_clean in {"?", "nzn", "nzn da", "nezinau"}:
        result["type"], result["parsed_text"] = "uncertain", "Dar nezino"
        return result
    if normalized_clean in {"a", "atostogos", "l", "liga", "n", "ne", "off"} or "negaliu" in normalized:
        result["type"], result["parsed_text"] = "unavailable", "Negali dirbti"
        return result
    if "atsidarym" in normalized:
        result["type"], result["parsed_text"] = "available", "Gali nuo atsidarymo"
        return result
    if "iki" in normalized and "nuo" in normalized:
        parts = re.split(r"\s*/\s*|\s+ir\s+", normalized)
        for part in parts:
            if "iki" in part:
                result["until_time"] = extract_time(part)
            if "nuo" in part:
                result["second_from_time"] = extract_time(part)
        result["type"] = "split"
        result["parsed_text"] = f"Gali iki {result['until_time']} ir nuo {result['second_from_time']}" if result["until_time"] and result["second_from_time"] else "Padalintas laikas"
        return result
    time_values = extract_all_times(normalized)
    if len(time_values) >= 2 and "-" in normalized and "iki" not in normalized and "nuo" not in normalized:
        result["type"], result["from_time"], result["until_time"] = "time_range", time_values[0], time_values[1]
        result["parsed_text"] = f"Gali nuo {time_values[0]} iki {time_values[1]}"
        return result
    if "nuo" in normalized:
        result["type"], result["from_time"] = "from_time", extract_time(normalized)
        result["parsed_text"] = f"Gali nuo {result['from_time']}" if result["from_time"] else "Gali nuo veliau"
        return result
    if "iki" in normalized:
        result["type"], result["until_time"] = "until_time", extract_time(normalized)
        result["parsed_text"] = f"Gali iki {result['until_time']}" if result["until_time"] else "Gali iki anksciau"
        return result
    if time_values:
        result["type"], result["from_time"], result["parsed_text"] = "from_time", time_values[0], f"Gali nuo {time_values[0]}"
        return result
    if "galiu" in normalized or result["preference"] in {"morning", "evening"}:
        result["type"] = "available"
        result["parsed_text"] = "Gali visa diena"
    return result


def get_ai_status():
    enabled = bool(os.getenv("OPENAI_API_KEY"))
    status = f"AI suggestions enabled ({DEFAULT_AI_MODEL})" if enabled else "Rules engine active"
    return {"enabled": enabled, "status": status}


def extract_response_text(response_json):
    output_text = response_json.get("output_text")
    if isinstance(output_text, str) and output_text.strip():
        return output_text.strip()
    parts = []
    for item in response_json.get("output", []) if isinstance(response_json.get("output"), list) else []:
        if item.get("type") != "message":
            continue
        for content_item in item.get("content", []):
            if content_item.get("type") == "output_text":
                parts.append(content_item.get("text", ""))
    return "\n".join(part for part in parts if part).strip()


def call_openai_json(prompt, schema_hint):
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        return None, "OPENAI_API_KEY nerastas"
    body = {
        "model": DEFAULT_AI_MODEL,
        "input": [
            {"role": "system", "content": f"Reply with valid JSON only. Follow this shape: {schema_hint}"},
            {"role": "user", "content": prompt},
        ],
    }
    request_obj = urllib.request.Request(
        "https://api.openai.com/v1/responses",
        data=json.dumps(body).encode("utf-8"),
        headers={"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(request_obj, timeout=20) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        try:
            details = exc.read().decode("utf-8")
        except Exception:
            details = str(exc)
        return None, f"AI klaida: {details[:200]}"
    except Exception as exc:
        return None, f"AI klaida: {exc}"
    raw_text = extract_response_text(payload)
    if not raw_text:
        return None, "AI negrazino teksto"
    try:
        return json.loads(raw_text), None
    except json.JSONDecodeError:
        for pattern in (r"\{.*\}", r"\[.*\]"):
            match = re.search(pattern, raw_text, re.DOTALL)
            if match:
                try:
                    return json.loads(match.group(0)), None
                except json.JSONDecodeError:
                    pass
        return None, "AI grazino ne JSON"


def normalize_ai_availability_result(original, payload):
    if not isinstance(payload, dict):
        return None
    result = build_unknown_parse_result(original)
    result["parser_source"] = "ai"
    result["confidence"] = safe_float(payload.get("confidence"), 0.0) or None
    result["ai_notes"] = str(payload.get("notes") or "").strip() or None
    parsed_type = str(payload.get("type") or "unknown").strip().lower()
    if parsed_type not in {"available", "unavailable", "uncertain", "from_time", "until_time", "time_range", "split", "unknown"}:
        parsed_type = "unknown"
    result["type"] = parsed_type
    for field_name in ("from_time", "until_time", "second_from_time"):
        value = payload.get(field_name)
        if value is not None and is_valid_time(str(value).strip()):
            result[field_name] = str(value).strip()
    preference = str(payload.get("preference") or "").strip().lower()
    if preference in {"morning", "evening"}:
        result["preference"] = preference
    parsed_text = str(payload.get("parsed_text") or "").strip()
    if parsed_text:
        result["parsed_text"] = parsed_text
    return result


def maybe_parse_availability_with_ai(line, fallback):
    if fallback["type"] not in {"unknown", "uncertain"}:
        return fallback
    prompt = (
        "Parse this employee availability note for one day and return JSON only. "
        "Use type from: available, unavailable, uncertain, from_time, until_time, time_range, split, unknown. "
        "Return HH:MM times. Text: " + line
    )
    schema_hint = '{"type":"unknown","from_time":null,"until_time":null,"second_from_time":null,"preference":null,"parsed_text":"","confidence":0.0,"notes":""}'
    payload, error = call_openai_json(prompt, schema_hint)
    normalized = normalize_ai_availability_result(line, payload)
    if normalized:
        if error:
            normalized["ai_notes"] = error
        return normalized
    fallback["ai_notes"] = error
    return fallback


def parse_single_line(line):
    return maybe_parse_availability_with_ai(line, parse_single_line_rule_based(line))


def check_one_shift_fit(parsed_item, shift_start, shift_end):
    shift_start_min = time_to_minutes(shift_start)
    shift_end_min = time_to_minutes(shift_end)
    result = {"ok": False, "reason": "Neatpazinta", "assigned_start": shift_start, "assigned_end": shift_end, "is_shortened": False}
    def valid_short_shift(start_min, end_min):
        return end_min - start_min >= MIN_SHORT_SHIFT_MINUTES
    if parsed_item["type"] == "available":
        result["ok"], result["reason"] = True, parsed_item["parsed_text"]
        return result
    if parsed_item["type"] in {"unavailable", "uncertain"}:
        result["reason"] = parsed_item["parsed_text"]
        return result
    if parsed_item["type"] == "from_time" and parsed_item["from_time"]:
        available_from = time_to_minutes(parsed_item["from_time"])
        if available_from <= shift_start_min:
            result["ok"], result["reason"] = True, f"Gali nuo {parsed_item['from_time']}"
            return result
        if shift_start_min < available_from < shift_end_min and valid_short_shift(available_from, shift_end_min):
            result.update({"ok": True, "reason": f"Sutrumpinta pamaina nuo {parsed_item['from_time']}", "assigned_start": minutes_to_time(available_from), "is_shortened": True})
        return result
    if parsed_item["type"] == "until_time" and parsed_item["until_time"]:
        available_until = time_to_minutes(parsed_item["until_time"])
        if available_until >= shift_end_min:
            result["ok"], result["reason"] = True, f"Gali iki {parsed_item['until_time']}"
            return result
        if shift_start_min < available_until < shift_end_min and valid_short_shift(shift_start_min, available_until):
            result.update({"ok": True, "reason": f"Sutrumpinta pamaina iki {parsed_item['until_time']}", "assigned_end": minutes_to_time(available_until), "is_shortened": True})
        return result
    if parsed_item["type"] == "time_range" and parsed_item["from_time"] and parsed_item["until_time"]:
        overlap_start = max(shift_start_min, time_to_minutes(parsed_item["from_time"]))
        overlap_end = min(shift_end_min, time_to_minutes(parsed_item["until_time"]))
        if overlap_start == shift_start_min and overlap_end == shift_end_min:
            result["ok"], result["reason"] = True, parsed_item["parsed_text"]
            return result
        if valid_short_shift(overlap_start, overlap_end):
            result.update({"ok": True, "reason": f"Sutrumpinta pamaina {minutes_to_time(overlap_start)}-{minutes_to_time(overlap_end)}", "assigned_start": minutes_to_time(overlap_start), "assigned_end": minutes_to_time(overlap_end), "is_shortened": True})
        return result
    if parsed_item["type"] == "split" and parsed_item["until_time"] and parsed_item["second_from_time"]:
        first_end = time_to_minutes(parsed_item["until_time"])
        second_start = time_to_minutes(parsed_item["second_from_time"])
        if shift_end_min <= first_end:
            result["ok"], result["reason"] = True, f"Gali iki {parsed_item['until_time']}"
            return result
        if second_start <= shift_start_min:
            result["ok"], result["reason"] = True, f"Gali nuo {parsed_item['second_from_time']}"
            return result
        if shift_start_min < first_end < shift_end_min and valid_short_shift(shift_start_min, first_end):
            result.update({"ok": True, "reason": f"Sutrumpinta pamaina iki {parsed_item['until_time']}", "assigned_end": minutes_to_time(first_end), "is_shortened": True})
            return result
        if shift_start_min < second_start < shift_end_min and valid_short_shift(second_start, shift_end_min):
            result.update({"ok": True, "reason": f"Sutrumpinta pamaina nuo {parsed_item['second_from_time']}", "assigned_start": minutes_to_time(second_start), "is_shortened": True})
        return result
    return result


def build_shift_fit_for_day(parsed_item, shifts):
    fit = []
    for shift in shifts:
        fit_result = check_one_shift_fit(parsed_item, shift["start"], shift["end"])
        fit.append({"label": shift["label"], "start": shift["start"], "end": shift["end"], "ok": fit_result["ok"], "reason": fit_result["reason"], "assigned_start": fit_result["assigned_start"], "assigned_end": fit_result["assigned_end"], "is_shortened": fit_result["is_shortened"]})
    return fit


def build_requested_shifts(location_id, year, month, day, demand_value):
    day_rules = get_day_rules_for_weekday(location_id, calendar.weekday(year, month, day))
    full_templates, half_templates = split_template_shifts(day_rules)
    full_count = int(demand_value)
    needs_half = abs(demand_value - full_count - 0.5) < 0.01
    warnings, shifts = [], []
    if full_count > len(full_templates):
        if full_templates:
            override_source = max(
                full_templates,
                key=lambda item: (time_to_minutes(item["end"]), time_to_minutes(item["start"])),
            )
            warnings.append(
                f"Poreikis {format_demand_value(demand_value)} virsija pilnu pamainu sablona ({len(full_templates)}). "
                f"Papildomos pamainos kuriamos kaip override pagal {override_source['start']}-{override_source['end']}."
            )
        else:
            warnings.append(f"Poreikis {format_demand_value(demand_value)} virsija pilnu pamainu sablona (0).")
    connected_half_shift = build_connected_half_shift(day_rules) if needs_half else None

    if full_count == 1 and needs_half and connected_half_shift:
        connected_full_shift = build_standard_opening_full_shift(day_rules, year, month, day)
        if connected_full_shift:
            connected_full_shift.update({"role": "full", "slot_kind": "Pilna", "label": "Pilna 1"})
            connected_full_shift = adjust_shift_for_special_opening(year, month, day, connected_full_shift, 0)
            shifts.append(connected_full_shift)
        half_shift = deepcopy(connected_half_shift)
        half_shift.update({"role": "half", "slot_kind": "Puse", "label": "Puse"})
        shifts.append(half_shift)
    else:
        for index in range(full_count):
            shift, is_override = get_full_shift_for_requested_index(full_templates, index)
            if not shift:
                continue
            if index == 0 and needs_half and connected_half_shift:
                standard_shift = build_standard_opening_full_shift(day_rules, year, month, day)
                if standard_shift:
                    shift = standard_shift
                    is_override = False
            shift.update({"role": "full", "slot_kind": "Pilna+" if is_override else "Pilna", "label": f"Pilna {index + 1}", "template_override": is_override})
            shift = adjust_shift_for_special_opening(year, month, day, shift, index)
            shifts.append(shift)
        if needs_half:
            if connected_half_shift:
                half_shift = deepcopy(connected_half_shift)
                half_shift.update({"role": "half", "slot_kind": "Puse", "label": "Puse"})
                shifts.append(half_shift)
            elif half_templates:
                shift = deepcopy(half_templates[0])
                shift.update({"role": "half", "slot_kind": "Puse", "label": "Puse"})
                shifts.append(shift)
            else:
                warnings.append("Pusinei pamainai nerastas sablonas.")
    shifts.sort(key=lambda item: (time_to_minutes(item["start"]), time_to_minutes(item["end"])))
    return shifts, warnings


def parse_worker_availability(lines, year, month, location_id, demand_map):
    parsed_days = []
    for day in range(1, get_days_in_month(year, month) + 1):
        original_line = lines[day - 1] if day - 1 < len(lines) else ""
        parsed = parse_single_line(original_line) if original_line else build_unknown_parse_result("", "Tuscia eilute")
        shifts, _ = build_requested_shifts(location_id, year, month, day, demand_map[day])
        parsed.update({"day": day, "day_type": get_day_type(year, month, day)[0], "weekday_name": weekday_name_from_index(calendar.weekday(year, month, day)), "shift_fit": build_shift_fit_for_day(parsed, shifts), "has_input": bool(original_line)})
        parsed_days.append(parsed)
    return parsed_days


def build_worker_runtime(worker, schedule_settings, location_id, demand_map):
    availability_lines = parse_availability_lines(worker.get("availability_raw", ""))
    worker_runtime = worker.copy()
    worker_runtime["availability_lines"] = availability_lines
    worker_runtime["parsed_availability"] = parse_worker_availability(availability_lines, schedule_settings["year"], schedule_settings["month"], location_id, demand_map)
    worker_runtime["day_count"] = len(availability_lines)
    worker_runtime["filled_day_count"] = sum(bool(line) for line in availability_lines)
    return worker_runtime


def calculate_targets(valid_workers, full_time_hours):
    return {
        index: etatas_to_month_hours(worker["etatas"], full_time_hours)
        for index, worker in enumerate(valid_workers)
    }


def get_assigned_shift_hours(shift, fit_info):
    return shift_length_hours(fit_info["assigned_start"], fit_info["assigned_end"]) if fit_info["is_shortened"] else shift.get("hours", shift_length_hours(shift["start"], shift["end"]))


def would_exceed_consecutive_limit(assigned_days, day, max_streak=MAX_CONSECUTIVE_DAYS):
    streak = 1
    previous_day = day - 1
    while previous_day in assigned_days:
        streak += 1
        previous_day -= 1
    next_day = day + 1
    while next_day in assigned_days:
        streak += 1
        next_day += 1
    return streak > max_streak


def would_break_rest_gap(assignments_by_day, day, assigned_start, assigned_end, min_rest_hours=MIN_REST_HOURS):
    start_min, end_min = time_to_minutes(assigned_start), time_to_minutes(assigned_end)
    previous_day_assignment = assignments_by_day.get(day - 1)
    if previous_day_assignment and (24 * 60 - previous_day_assignment["end"]) + start_min < min_rest_hours * 60:
        return True
    next_day_assignment = assignments_by_day.get(day + 1)
    if next_day_assignment and (24 * 60 - end_min) + next_day_assignment["start"] < min_rest_hours * 60:
        return True
    return False


def get_soft_turnaround_penalty(assignments_by_day, day, assigned_start, assigned_end, preferred_rest_hours=PREFERRED_REST_HOURS):
    start_min, end_min = time_to_minutes(assigned_start), time_to_minutes(assigned_end)
    preferred_minutes = preferred_rest_hours * 60
    penalty = 0.0

    previous_day_assignment = assignments_by_day.get(day - 1)
    if previous_day_assignment:
        previous_gap = (24 * 60 - previous_day_assignment["end"]) + start_min
        penalty += max(0, preferred_minutes - previous_gap) / 60

    next_day_assignment = assignments_by_day.get(day + 1)
    if next_day_assignment:
        next_gap = (24 * 60 - end_min) + next_day_assignment["start"]
        penalty += max(0, preferred_minutes - next_gap) / 60

    return penalty * SOFT_TURNAROUND_PENALTY_PER_HOUR


def get_adjacent_start_bonus(assignments_by_day, day, assigned_start):
    start_min = time_to_minutes(assigned_start)
    matches = 0
    for adjacent_day in (day - 1, day + 1):
        adjacent_assignment = assignments_by_day.get(adjacent_day)
        if adjacent_assignment and abs(adjacent_assignment["start"] - start_min) <= ADJACENT_START_TOLERANCE_MINUTES:
            matches += 1
    return matches * ADJACENT_START_BONUS


def get_processing_day_order(year, month, demand_map):
    def sort_key(day):
        day_type, _ = get_day_type(year, month, day)
        weekend_priority = 0 if day_type == "weekend" else 1
        return (-demand_map[day], weekend_priority, day)
    return sorted(range(1, get_days_in_month(year, month) + 1), key=sort_key)


def get_day_closing_end_minutes(shifts):
    return max(time_to_minutes(shift["end"]) for shift in shifts) if shifts else None


def is_closing_assignment(assigned_end, closing_end_minutes):
    return closing_end_minutes is not None and time_to_minutes(assigned_end) == closing_end_minutes


def assignment_time_to_minutes(shift_time):
    start_str, end_str = shift_time.split("-")
    return time_to_minutes(start_str), time_to_minutes(end_str)


def parse_assignment_interval(shift_time):
    match = TIME_RANGE_FULL_RE.fullmatch(str(shift_time or ""))
    return interval_from_match(match) if match else None


def can_cover_interval(parsed_day, start_time, end_time):
    fit = check_one_shift_fit(parsed_day, start_time, end_time)
    return fit["ok"] and fit["assigned_start"] == start_time and fit["assigned_end"] == end_time


def repair_gap_with_availability(assignments, worker_day_map):
    usable = [assignment for assignment in assignments if assignment["worker_name"]]
    if len(usable) < 2:
        return False, None

    usable_sorted = sorted(usable, key=lambda assignment: assignment_time_to_minutes(assignment["shift_time"])[0])

    for index in range(len(usable_sorted) - 1):
        current = usable_sorted[index]
        following = usable_sorted[index + 1]
        current_start, current_end = assignment_time_to_minutes(current["shift_time"])
        next_start, next_end = assignment_time_to_minutes(following["shift_time"])

        if current_end >= next_start:
            continue

        gap_start = minutes_to_time(current_end)
        gap_end = minutes_to_time(next_start)

        current_day = worker_day_map.get(current["worker_index"])
        if not current.get("locked") and current_day and can_cover_interval(current_day, minutes_to_time(current_start), gap_end):
            current["shift_time"] = f"{minutes_to_time(current_start)}-{gap_end}"
            return True, f"Tarpas uzdengtas prailginant {current['worker_name']} pamaina iki {gap_end} (sulauzo bazini sablona)."

        following_day = worker_day_map.get(following["worker_index"])
        if not following.get("locked") and following_day and can_cover_interval(following_day, gap_start, minutes_to_time(next_end)):
            following["shift_time"] = f"{gap_start}-{minutes_to_time(next_end)}"
            return True, f"Tarpas uzdengtas paankstinant {following['worker_name']} pamaina nuo {gap_start} (sulauzo bazini sablona)."

        if not current.get("locked") and not following.get("locked") and current_day and following_day:
            for bridge_minute in range(current_end + 30, next_start + 1, 30):
                bridge_time = minutes_to_time(bridge_minute)
                current_can_extend = can_cover_interval(
                    current_day,
                    minutes_to_time(current_start),
                    bridge_time,
                )
                following_can_pull = can_cover_interval(
                    following_day,
                    bridge_time,
                    minutes_to_time(next_end),
                )

                if current_can_extend and following_can_pull:
                    current["shift_time"] = f"{minutes_to_time(current_start)}-{bridge_time}"
                    following["shift_time"] = f"{bridge_time}-{minutes_to_time(next_end)}"
                    return (
                        True,
                        f"Tarpas uzdengtas pastumiant abi pamainas iki {bridge_time} (sulauzo bazini sablona).",
                    )

    return False, None


def has_gap_between_assignments(assignments):
    usable = [assignment for assignment in assignments if assignment["worker_name"]]
    if len(usable) < 2:
        return False
    usable_sorted = sorted(usable, key=lambda assignment: assignment_time_to_minutes(assignment["shift_time"])[0])
    return any(assignment_time_to_minutes(usable_sorted[index]["shift_time"])[1] < assignment_time_to_minutes(usable_sorted[index + 1]["shift_time"])[0] for index in range(len(usable_sorted) - 1))


def sort_workers_for_display(worker_list):
    return sorted(worker_list, key=lambda worker: (-safe_float(worker["etatas"]), worker["name"].lower()))


def get_shift_priority_score(worker_index, parsed_day, fit_info, shift, projected_hours, target_hours, assigned_hours, assigned_counts, weekend_counts, closing_counts, day_type, is_closing, turnaround_penalty=0.0, adjacent_start_bonus=0.0):
    score = assigned_hours[worker_index] - target_hours
    if shift["role"] == "full" and parsed_day.get("preference") == "morning" and shift["start"] <= "11:00":
        score -= 0.25
    if parsed_day.get("preference") == "evening" and shift["end"] >= "21:00":
        score -= 0.2
    if fit_info["is_shortened"]:
        score += 0.2
    if projected_hours > target_hours:
        score += 5 + (projected_hours - target_hours)
    if day_type == "weekend":
        score += weekend_counts[worker_index] * 0.7
    if is_closing:
        score += closing_counts[worker_index] * 0.8
    score += assigned_counts[worker_index] * 0.05
    score += turnaround_penalty
    score -= adjacent_start_bonus
    return score


def blocks_closing_shift(parsed_day, is_closing):
    return is_closing and parsed_day.get("preference") == "morning"


def summarize_rejection_reasons(reason_counts):
    if not reason_counts:
        return "nerasta tinkamu kandidatu"

    priority = [
        "jau priskirtas kita pamaina ta diena",
        "negali dirbti sios pamainos",
        "virsytu 5 dienas is eiles",
        "per mazai poilsio tarp pamainu",
        "ryto pageidavimas blokuoja uzdaryma",
        "virsytu valandu limita",
        "nera ivestu galimybiu tai dienai",
    ]

    best_reason = None
    best_count = -1
    for reason in priority:
        count = reason_counts.get(reason, 0)
        if count > best_count:
            best_reason = reason
            best_count = count

    if best_count <= 0:
        best_reason, best_count = max(reason_counts.items(), key=lambda item: item[1])

    return f"{best_reason} ({best_count})"


def append_unique_warning(day_record, warning):
    if warning not in day_record["warnings"]:
        day_record["warnings"].append(warning)


def build_existing_schedule_map(existing_schedule, days_in_month):
    existing_by_day = {}
    for raw_day in existing_schedule if isinstance(existing_schedule, list) else []:
        if not isinstance(raw_day, dict):
            continue
        day = sanitize_int(raw_day.get("day"), 0, 1, days_in_month)
        if day:
            existing_by_day[day] = raw_day
    return existing_by_day


def build_effective_template_shift(template_shift, interval, existing_assignment):
    effective_shift = deepcopy(template_shift)
    if interval != (template_shift["start"], template_shift["end"]):
        hours = shift_length_hours(interval[0], interval[1])
        effective_shift.update({
            "start": interval[0],
            "end": interval[1],
            "hours": hours,
            "role": "full" if hours >= FULL_SHIFT_MIN_HOURS else "half",
        })
    effective_shift["label"] = str(existing_assignment.get("shift_label") or effective_shift["label"])
    effective_shift["slot_kind"] = str(existing_assignment.get("slot_kind") or effective_shift["slot_kind"])
    return effective_shift


def build_schedule_blueprint(existing_schedule, schedule_settings, location_id, demand_map):
    year, month = schedule_settings["year"], schedule_settings["month"]
    days_in_month = get_days_in_month(year, month)
    existing_by_day = build_existing_schedule_map(existing_schedule, days_in_month)
    day_records = {}

    for day in range(1, days_in_month + 1):
        day_type, weekday_index = get_day_type(year, month, day)
        shifts, day_warnings = build_requested_shifts(location_id, year, month, day, demand_map[day])
        existing_day = existing_by_day.get(day, {})
        existing_assignments = existing_day.get("assignments", []) if isinstance(existing_day.get("assignments"), list) else []
        assignments = []
        day_record = {
            "day": day,
            "weekday_name": weekday_name_from_index(weekday_index),
            "day_type": day_type,
            "demand_value": demand_map[day],
            "demand_label": format_demand_value(demand_map[day]),
            "requested_hours": int(demand_map[day] * 8),
            "assignments": assignments,
            "warnings": list(day_warnings),
            "closing_end_minutes": get_day_closing_end_minutes(shifts),
        }

        for shift_index, shift in enumerate(shifts):
            existing_assignment = existing_assignments[shift_index] if shift_index < len(existing_assignments) and isinstance(existing_assignments[shift_index], dict) else {}
            interval = parse_assignment_interval(existing_assignment.get("shift_time"))
            if not interval:
                interval = shift["start"], shift["end"]
                if existing_assignment.get("worker_name") or existing_assignment.get("worker_id"):
                    append_unique_warning(day_record, f"Rankines pamainos {shift['label']} laikas buvo netinkamas; paliktas sablono laikas.")
            if existing_assignment.get("time_warning"):
                append_unique_warning(day_record, str(existing_assignment["time_warning"]))
            worker_name = str(existing_assignment.get("worker_name") or "").strip() or None
            worker_id = str(existing_assignment.get("worker_id") or "").strip() or None
            effective_shift = build_effective_template_shift(shift, interval, existing_assignment)
            assignments.append({
                "shift_label": effective_shift["label"],
                "slot_kind": effective_shift["slot_kind"],
                "shift_time": f"{interval[0]}-{interval[1]}",
                "worker_name": worker_name,
                "worker_id": worker_id,
                "worker_index": None,
                "locked": bool(worker_name or worker_id),
                "template_shift": effective_shift,
            })

        for extra_index, existing_assignment in enumerate(existing_assignments[len(shifts):], start=len(shifts) + 1):
            if not isinstance(existing_assignment, dict):
                continue
            worker_name = str(existing_assignment.get("worker_name") or "").strip() or None
            worker_id = str(existing_assignment.get("worker_id") or "").strip() or None
            if not worker_name and not worker_id:
                continue
            interval = parse_assignment_interval(existing_assignment.get("shift_time"))
            if not interval:
                append_unique_warning(day_record, f"Papildoma rankine pamaina {extra_index} praleista, nes jos laikas netinkamas.")
                continue
            hours = shift_length_hours(interval[0], interval[1])
            extra_shift = {
                "label": str(existing_assignment.get("shift_label") or f"Rankine {extra_index}"),
                "slot_kind": str(existing_assignment.get("slot_kind") or "Rankine"),
                "start": interval[0],
                "end": interval[1],
                "hours": hours,
                "role": "full" if hours >= FULL_SHIFT_MIN_HOURS else "half",
            }
            assignments.append({
                "shift_label": extra_shift["label"],
                "slot_kind": extra_shift["slot_kind"],
                "shift_time": f"{interval[0]}-{interval[1]}",
                "worker_name": worker_name,
                "worker_id": worker_id,
                "worker_index": None,
                "locked": True,
                "template_shift": extra_shift,
            })
            append_unique_warning(day_record, "Palikta papildoma rankine pamaina uz poreikio sablono ribu.")

        assignment_ends = [
            time_to_minutes(interval[1])
            for assignment in assignments
            if (interval := parse_assignment_interval(assignment["shift_time"]))
        ]
        if assignment_ends:
            day_record["closing_end_minutes"] = max(day_record["closing_end_minutes"] or 0, max(assignment_ends))
        if demand_map[day] == 0:
            append_unique_warning(day_record, "Lokacija nedirba")
        day_records[day] = day_record

    return day_records


def build_worker_assignment_lookup(valid_workers):
    workers_by_id = {str(worker.get("id")): index for index, worker in enumerate(valid_workers) if worker.get("id")}
    workers_by_name = {}
    for index, worker in enumerate(valid_workers):
        workers_by_name.setdefault(normalize_text(worker["name"]), []).append(index)
    return workers_by_id, workers_by_name


def resolve_assignment_worker_index(assignment, workers_by_id, workers_by_name):
    worker_id = str(assignment.get("worker_id") or "").strip()
    if worker_id and worker_id in workers_by_id:
        return workers_by_id[worker_id]
    worker_name = normalize_text(assignment.get("worker_name"))
    matches = workers_by_name.get(worker_name, []) if worker_name else []
    return matches[0] if len(matches) == 1 else None


def register_assignment_state(worker_index, day, day_type, assignment, closing_end_minutes, assigned_counts, assigned_hours, assigned_days, weekend_counts, closing_counts, assignments_by_worker_day):
    start_time, end_time = parse_assignment_interval(assignment["shift_time"])
    template_shift = assignment["template_shift"]
    if start_time == template_shift["start"] and end_time == template_shift["end"]:
        shift_hours = template_shift.get("hours", shift_length_hours(start_time, end_time))
    else:
        shift_hours = shift_length_hours(start_time, end_time)

    assigned_counts[worker_index] += 1
    assigned_hours[worker_index] += shift_hours
    assigned_days[worker_index].add(day)
    if day_type == "weekend":
        weekend_counts[worker_index] += 1
    if is_closing_assignment(end_time, closing_end_minutes):
        closing_counts[worker_index] += 1

    start_minutes, end_minutes = time_to_minutes(start_time), time_to_minutes(end_time)
    existing_window = assignments_by_worker_day[worker_index].get(day)
    if existing_window:
        existing_window["start"] = min(existing_window["start"], start_minutes)
        existing_window["end"] = max(existing_window["end"], end_minutes)
    else:
        assignments_by_worker_day[worker_index][day] = {"start": start_minutes, "end": end_minutes}


def add_locked_assignment_warnings(day_records, valid_workers, assignments_by_worker_day, assigned_days):
    for worker_index, worker in enumerate(valid_workers):
        prior_day = None
        streak = 0
        for day in sorted(assigned_days[worker_index]):
            streak = streak + 1 if prior_day is not None and day == prior_day + 1 else 1
            if streak > MAX_CONSECUTIVE_DAYS:
                append_unique_warning(day_records[day], f"Rankinis pasirinkimas: {worker['name']} dirba daugiau nei {MAX_CONSECUTIVE_DAYS} dienas is eiles.")
            prior_day = day

        worker_days = assignments_by_worker_day[worker_index]
        for day in sorted(worker_days):
            previous = worker_days.get(day - 1)
            if not previous:
                continue
            rest_minutes = (24 * 60 - previous["end"]) + worker_days[day]["start"]
            if rest_minutes < MIN_REST_HOURS * 60:
                append_unique_warning(day_records[day], f"Rankinis pasirinkimas: {worker['name']} turi maziau nei {MIN_REST_HOURS} val. poilsio.")


def generate_month_schedule(worker_list, schedule_settings, location_id, demand_map, existing_schedule=None, fill_open_slots=True):
    year, month = schedule_settings["year"], schedule_settings["month"]
    valid_workers = sort_workers_for_display(worker_list[:])
    targets = calculate_targets(valid_workers, schedule_settings["full_time_hours"])
    assigned_counts = {index: 0 for index in range(len(valid_workers))}
    assigned_hours = {index: 0.0 for index in range(len(valid_workers))}
    assigned_days = {index: set() for index in range(len(valid_workers))}
    weekend_counts = {index: 0 for index in range(len(valid_workers))}
    closing_counts = {index: 0 for index in range(len(valid_workers))}
    assignments_by_worker_day = {index: {} for index in range(len(valid_workers))}
    workers_by_id, workers_by_name = build_worker_assignment_lookup(valid_workers)
    day_records_by_day = build_schedule_blueprint(existing_schedule, schedule_settings, location_id, demand_map)

    for day, day_record in day_records_by_day.items():
        seen_workers = set()
        for assignment in day_record["assignments"]:
            if not assignment["locked"]:
                continue
            worker_index = resolve_assignment_worker_index(assignment, workers_by_id, workers_by_name)
            assignment["worker_index"] = worker_index
            if worker_index is None:
                append_unique_warning(day_record, f"Rankine pamaina palikta, bet darbuotojas nerastas: {assignment['worker_name'] or assignment['worker_id']}.")
                continue
            worker = valid_workers[worker_index]
            assignment["worker_id"] = worker.get("id")
            assignment["worker_name"] = worker["name"]
            if worker_index in seen_workers:
                append_unique_warning(day_record, f"Rankinis pasirinkimas: {worker['name']} turi daugiau nei viena pamaina ta pacia diena.")
            seen_workers.add(worker_index)
            parsed_day = worker["parsed_availability"][day - 1]
            start_time, end_time = parse_assignment_interval(assignment["shift_time"])
            if not can_cover_interval(parsed_day, start_time, end_time):
                append_unique_warning(day_record, f"Rankinis pasirinkimas neatitinka {worker['name']} galimybiu ({start_time}-{end_time}).")
            register_assignment_state(
                worker_index,
                day,
                day_record["day_type"],
                assignment,
                day_record["closing_end_minutes"],
                assigned_counts,
                assigned_hours,
                assigned_days,
                weekend_counts,
                closing_counts,
                assignments_by_worker_day,
            )

    add_locked_assignment_warnings(day_records_by_day, valid_workers, assignments_by_worker_day, assigned_days)

    for day in get_processing_day_order(year, month, demand_map):
        day_record = day_records_by_day[day]
        assignments_for_day = day_record["assignments"]
        assigned_today = {assignment["worker_index"] for assignment in assignments_for_day if assignment["locked"] and assignment["worker_index"] is not None}
        assigned_worker_days = {
            assignment["worker_index"]: valid_workers[assignment["worker_index"]]["parsed_availability"][day - 1]
            for assignment in assignments_for_day
            if assignment["worker_index"] is not None
        }

        for assignment in assignments_for_day:
            if assignment["locked"]:
                continue
            shift = assignment["template_shift"]
            if not fill_open_slots:
                continue

            candidates = []
            rejection_reasons = {}
            for worker_index, worker in enumerate(valid_workers):
                if worker_index in assigned_today:
                    rejection_reasons["jau priskirtas kita pamaina ta diena"] = rejection_reasons.get("jau priskirtas kita pamaina ta diena", 0) + 1
                    continue
                parsed_day = worker["parsed_availability"][day - 1]
                fit_info = check_one_shift_fit(parsed_day, shift["start"], shift["end"])
                if not fit_info["ok"]:
                    rejection_reasons["negali dirbti sios pamainos"] = rejection_reasons.get("negali dirbti sios pamainos", 0) + 1
                    continue
                if would_exceed_consecutive_limit(assigned_days[worker_index], day):
                    rejection_reasons["virsytu 5 dienas is eiles"] = rejection_reasons.get("virsytu 5 dienas is eiles", 0) + 1
                    continue
                if would_break_rest_gap(assignments_by_worker_day[worker_index], day, fit_info["assigned_start"], fit_info["assigned_end"]):
                    rejection_reasons["per mazai poilsio tarp pamainu"] = rejection_reasons.get("per mazai poilsio tarp pamainu", 0) + 1
                    continue
                shift_hours = get_assigned_shift_hours(shift, fit_info)
                projected_hours = assigned_hours[worker_index] + shift_hours
                if projected_hours > targets[worker_index] + 24:
                    rejection_reasons["virsytu valandu limita"] = rejection_reasons.get("virsytu valandu limita", 0) + 1
                    continue
                is_closing = is_closing_assignment(fit_info["assigned_end"], day_record["closing_end_minutes"])
                if blocks_closing_shift(parsed_day, is_closing):
                    rejection_reasons["ryto pageidavimas blokuoja uzdaryma"] = rejection_reasons.get("ryto pageidavimas blokuoja uzdaryma", 0) + 1
                    continue
                turnaround_penalty = get_soft_turnaround_penalty(assignments_by_worker_day[worker_index], day, fit_info["assigned_start"], fit_info["assigned_end"])
                adjacent_start_bonus = get_adjacent_start_bonus(assignments_by_worker_day[worker_index], day, fit_info["assigned_start"])
                score = get_shift_priority_score(worker_index, parsed_day, fit_info, shift, projected_hours, targets[worker_index], assigned_hours, assigned_counts, weekend_counts, closing_counts, day_record["day_type"], is_closing, turnaround_penalty, adjacent_start_bonus)
                candidates.append((score, assigned_counts[worker_index], worker_index, worker["name"], fit_info, shift_hours, is_closing))

            candidates.sort(key=lambda item: (item[0], item[1], item[3].lower()))
            if candidates:
                _, _, chosen_index, chosen_name, chosen_fit, chosen_shift_hours, chosen_is_closing = candidates[0]
                assigned_today.add(chosen_index)
                assigned_counts[chosen_index] += 1
                assigned_hours[chosen_index] += chosen_shift_hours
                assigned_days[chosen_index].add(day)
                if day_record["day_type"] == "weekend":
                    weekend_counts[chosen_index] += 1
                if chosen_is_closing:
                    closing_counts[chosen_index] += 1
                assignments_by_worker_day[chosen_index][day] = {"start": time_to_minutes(chosen_fit["assigned_start"]), "end": time_to_minutes(chosen_fit["assigned_end"])}
                assigned_worker_days[chosen_index] = valid_workers[chosen_index]["parsed_availability"][day - 1]
                assignment.update({
                    "shift_time": f"{chosen_fit['assigned_start']}-{chosen_fit['assigned_end']}",
                    "worker_name": chosen_name,
                    "worker_id": valid_workers[chosen_index].get("id"),
                    "worker_index": chosen_index,
                })
            else:
                reason_text = summarize_rejection_reasons(rejection_reasons)
                append_unique_warning(day_record, f"Nera darbuotojo {shift['label']} ({shift['start']}-{shift['end']}) del: {reason_text}")

        if fill_open_slots:
            gap_repaired, gap_warning = repair_gap_with_availability(assignments_for_day, assigned_worker_days)
            if gap_repaired and gap_warning:
                append_unique_warning(day_record, gap_warning)
        else:
            open_shift_count = sum(not assignment["worker_name"] for assignment in assignments_for_day)
            if open_shift_count:
                append_unique_warning(day_record, f"Nepriskirta pamainu: {open_shift_count}.")
        if has_gap_between_assignments(assignments_for_day):
            append_unique_warning(day_record, "Yra tarpas grafike - parduotuve liktu tuscia")

    generated_schedule = []
    for day in range(1, get_days_in_month(year, month) + 1):
        day_record = day_records_by_day[day]
        day_record.pop("closing_end_minutes", None)
        for assignment in day_record["assignments"]:
            assignment.pop("worker_index", None)
            assignment.pop("locked", None)
            assignment.pop("template_shift", None)
        generated_schedule.append(day_record)

    worker_summary = [{"name": worker["name"], "etatas": worker["etatas"], "assigned_shifts": assigned_counts[index], "assigned_hours": round(assigned_hours[index], 1), "target_hours": round(targets[index], 1), "hours_difference": round(assigned_hours[index] - targets[index], 1), "weekend_days": weekend_counts[index], "closing_shifts": closing_counts[index]} for index, worker in enumerate(valid_workers)]
    worker_summary = sorted(worker_summary, key=lambda worker: (-safe_float(worker["etatas"]), worker["name"].lower()))
    return generated_schedule, worker_summary


def build_rule_insights(generated_schedule, worker_summary):
    insights = []
    uncovered_days = [day for day in generated_schedule if any(not assignment["worker_name"] for assignment in day["assignments"])]
    if uncovered_days:
        insights.append(f"Truksta padengimo {len(uncovered_days)} dienu: " + ", ".join(str(day["day"]) for day in uncovered_days[:8]) + ".")
    overloaded = [worker for worker in worker_summary if worker["hours_difference"] > 8]
    if overloaded:
        top_worker = max(overloaded, key=lambda item: item["hours_difference"])
        insights.append(f"{top_worker['name']} yra labiausiai perkrautas: +{top_worker['hours_difference']} val.")
    underloaded = [worker for worker in worker_summary if worker["hours_difference"] < -8]
    if underloaded:
        top_worker = min(underloaded, key=lambda item: item["hours_difference"])
        insights.append(f"{top_worker['name']} labiausiai atsilieka nuo etato: {top_worker['hours_difference']} val.")
    weekend_heavy = [worker for worker in worker_summary if worker["weekend_days"] >= 4]
    if weekend_heavy:
        insights.append("Daug savaitgaliu tenka: " + ", ".join(worker["name"] for worker in weekend_heavy[:4]) + ".")
    return insights or ["Pagal taisykles grafikas atrodo subalansuotas."]


def build_ai_schedule_suggestions(location_name, generated_schedule, worker_summary):
    prompt = "Pasiulyk iki 5 trumpu praktisku patobulinimu lietuviskai siam grafikui. Grazink JSON. " + json.dumps({"location": location_name, "uncovered_days": [{"day": day["day"], "warnings": day["warnings"]} for day in generated_schedule if any(not assignment["worker_name"] for assignment in day["assignments"])], "worker_summary": worker_summary}, ensure_ascii=False)
    response_json, error = call_openai_json(prompt, '{"suggestions":["..."]}')
    if error:
        return None, error
    if isinstance(response_json, dict) and isinstance(response_json.get("suggestions"), list):
        cleaned = [str(item).strip() for item in response_json["suggestions"] if str(item).strip()]
        if cleaned:
            return cleaned[:5], None
    return None, "AI nepateike pasiulymu"


def build_schedule_insights(location, generated_schedule, worker_summary):
    items = build_rule_insights(generated_schedule, worker_summary)
    ai_status = get_ai_status()
    if not ai_status["enabled"]:
        return {"items": items, "ai_used": False, "ai_status": ai_status["status"]}
    ai_items, ai_error = build_ai_schedule_suggestions(location["name"], generated_schedule, worker_summary)
    if ai_items:
        items.extend(ai_items)
    return {"items": items, "ai_used": bool(ai_items), "ai_status": ai_error or ai_status["status"]}


def autosize_worksheet_columns(worksheet):
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 40)


EXCEL_MONTH_GENITIVE_NAMES = {
    1: "sausio",
    2: "vasario",
    3: "kovo",
    4: "balandžio",
    5: "gegužės",
    6: "birželio",
    7: "liepos",
    8: "rugpjūčio",
    9: "rugsėjo",
    10: "spalio",
    11: "lapkričio",
    12: "gruodžio",
}
EXCEL_WEEKDAY_LETTERS = ["P", "A", "T", "K", "P", "Š", "S"]


class ScheduleImportError(ValueError):
    def __init__(self, code):
        super().__init__(code)
        self.code = code


def parse_import_date_value(value, default_year):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None

    for date_format in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            pass

    normalized = normalize_text(text)
    for month, month_name in EXCEL_MONTH_GENITIVE_NAMES.items():
        aliases = {normalize_text(month_name), normalize_text(MONTH_NAMES[month])}
        for alias in aliases:
            match = re.fullmatch(rf"{re.escape(alias)}\s+(\d{{1,2}})(?:\s+d\.?)?(?:\s+(\d{{4}}))?", normalized)
            if not match:
                continue
            day_number = int(match.group(1))
            year = int(match.group(2) or default_year)
            try:
                return date(year, month, day_number)
            except ValueError:
                return None
    return None


def parse_import_demand_value(value):
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        parsed = float(value)
    else:
        text = str(value or "").strip()
        if not re.fullmatch(r"\d(?:[.,]\d+)?", text):
            return None
        parsed = safe_float(text, -1)
    if not 0 <= parsed <= 5 or abs(parsed * 2 - round(parsed * 2)) > 0.01:
        return None
    return round(parsed * 2) / 2


def expected_import_slot_count(demand_value):
    full_count = int(demand_value)
    return min(MAX_IMPORT_SHIFT_COLUMNS, full_count + (1 if demand_value - full_count >= 0.49 else 0))


def resolve_import_worker(candidate_name, workers):
    normalized_candidate = normalize_text(candidate_name)
    if not normalized_candidate:
        return None

    exact_matches = [worker for worker in workers if normalize_text(worker.get("name")) == normalized_candidate]
    if len(exact_matches) == 1:
        return exact_matches[0]

    mentioned = []
    for worker in workers:
        normalized_worker = normalize_text(worker.get("name"))
        if normalized_worker and re.search(rf"(?<!\w){re.escape(normalized_worker)}(?!\w)", normalized_candidate):
            mentioned.append(worker)
    if len(mentioned) == 1:
        return mentioned[0]
    if len(mentioned) > 1:
        return None

    first_token = normalized_candidate.split()[0]
    first_name_matches = [
        worker for worker in workers
        if normalize_text(worker.get("name")).split()[0] == first_token
    ]
    return first_name_matches[0] if len(first_name_matches) == 1 else None


def parse_import_assignment_cell(value, workers, slot_index):
    text = str(value or "").strip()
    if not text or text in {"-", "–", "—"}:
        return None, False, False
    match = TIME_RANGE_SEARCH_RE.search(text)
    interval = interval_from_match(match) if match else None
    if not interval:
        return None, True, False

    candidate_name = re.sub(r"\s+", " ", text[:match.start()].strip(" -–—"))
    worker = resolve_import_worker(candidate_name, workers) if candidate_name else None
    hours = shift_length_hours(interval[0], interval[1])
    assignment = {
        "shift_label": f"Importuota {slot_index + 1}",
        "slot_kind": "Pilna" if hours >= FULL_SHIFT_MIN_HOURS else "Puse",
        "shift_time": f"{interval[0]}-{interval[1]}",
        "worker_id": worker.get("id") if worker else None,
        "worker_name": worker.get("name") if worker else (candidate_name or None),
    }
    return assignment, False, bool(candidate_name and not worker)


def find_import_row_date(row, default_year):
    for column_index, value in enumerate(row[:6]):
        parsed_date = parse_import_date_value(value, default_year)
        if parsed_date:
            return column_index, parsed_date
    return None, None


def find_import_row_demand(row, date_column):
    for value in row[:date_column]:
        demand_value = parse_import_demand_value(value)
        if demand_value is not None:
            return demand_value
    return None


def infer_blank_import_times(imported_days, year, month):
    sources = []
    for day_number, day_data in imported_days.items():
        weekday_index = calendar.weekday(year, month, day_number)
        for slot_index, assignment in enumerate(day_data["assignments"]):
            if not isinstance(assignment, dict):
                continue
            interval = parse_assignment_interval(assignment.get("shift_time"))
            if interval:
                sources.append((day_number, weekday_index, slot_index, interval))

    inferred_count = 0
    for day_number, day_data in imported_days.items():
        weekday_index = calendar.weekday(year, month, day_number)
        for slot_index, assignment in enumerate(day_data["assignments"]):
            if isinstance(assignment, dict):
                continue
            candidates = [source for source in sources if source[2] == slot_index]
            if not candidates:
                continue

            def candidate_key(source):
                source_day, source_weekday, _, _ = source
                if source_weekday == weekday_index and source_day <= day_number:
                    group = 0
                elif source_weekday == weekday_index:
                    group = 1
                elif source_day <= day_number:
                    group = 2
                else:
                    group = 3
                return group, abs(day_number - source_day), -source_day

            _, _, _, interval = min(candidates, key=candidate_key)
            hours = shift_length_hours(interval[0], interval[1])
            day_data["assignments"][slot_index] = {
                "shift_label": f"Importuota {slot_index + 1}",
                "slot_kind": "Pilna" if hours >= FULL_SHIFT_MIN_HOURS else "Puse",
                "shift_time": f"{interval[0]}-{interval[1]}",
                "worker_id": None,
                "worker_name": None,
            }
            inferred_count += 1
    return inferred_count


def parse_partial_schedule_rows(rows, location, location_id):
    settings = location["schedule_settings"]
    year, month = settings["year"], settings["month"]
    current_demand, _ = get_demand_context(location, location_id)
    imported_days = {}
    demand_updates = {}
    warning_count = 0
    unknown_worker_count = 0
    assigned_count = 0

    for row in rows[:MAX_IMPORT_ROWS]:
        values = list(row[:MAX_IMPORT_COLUMNS])
        date_column, row_date = find_import_row_date(values, year)
        if not row_date or row_date.year != year or row_date.month != month:
            continue

        day_number = row_date.day
        demand_value = find_import_row_demand(values, date_column)
        if demand_value is None:
            demand_value = current_demand[day_number]
        else:
            demand_updates[day_number] = demand_value
        slot_count = expected_import_slot_count(demand_value)
        assignment_cells = values[date_column + 1:date_column + 1 + MAX_IMPORT_SHIFT_COLUMNS]
        assignments = []

        for slot_index in range(slot_count):
            cell_value = assignment_cells[slot_index] if slot_index < len(assignment_cells) else None
            assignment, invalid, unknown = parse_import_assignment_cell(cell_value, location["workers"], slot_index)
            assignments.append(assignment)
            warning_count += int(invalid)
            unknown_worker_count += int(unknown)
            assigned_count += int(bool(assignment and assignment.get("worker_name")))

        for slot_index in range(slot_count, min(len(assignment_cells), MAX_IMPORT_SHIFT_COLUMNS)):
            assignment, _, unknown = parse_import_assignment_cell(assignment_cells[slot_index], location["workers"], slot_index)
            if not assignment:
                continue
            assignments.append(assignment)
            unknown_worker_count += int(unknown)
            assigned_count += int(bool(assignment.get("worker_name")))

        imported_days[day_number] = {"day": day_number, "assignments": assignments}

    if not imported_days:
        raise ScheduleImportError("no_rows")

    inferred_time_count = infer_blank_import_times(imported_days, year, month)
    return {
        "schedule": [imported_days[day] for day in sorted(imported_days)],
        "demand_updates": demand_updates,
        "day_count": len(imported_days),
        "assigned_count": assigned_count,
        "inferred_time_count": inferred_time_count,
        "unknown_worker_count": unknown_worker_count,
        "warning_count": warning_count,
    }


def get_import_sheet_candidates(location, location_id):
    return [location.get("name"), get_location_rule(location_id).get("name"), location_id.replace("-", " ")]


def select_import_worksheet(workbook, requested_sheet_name, location, location_id):
    worksheets = list(workbook.worksheets)
    if not worksheets:
        raise ScheduleImportError("unreadable")

    normalized_titles = {normalize_text(sheet.title): sheet for sheet in worksheets}
    if requested_sheet_name:
        requested = normalized_titles.get(normalize_text(requested_sheet_name))
        if not requested:
            raise ScheduleImportError("sheet_not_found")
        return requested

    candidates = [normalize_text(value) for value in get_import_sheet_candidates(location, location_id) if value]
    for candidate in candidates:
        if candidate in normalized_titles:
            return normalized_titles[candidate]

    if "grafikas" in normalized_titles:
        return normalized_titles["grafikas"]
    if len(worksheets) == 1:
        return worksheets[0]

    ranked = []
    for sheet in worksheets:
        normalized_title = normalize_text(sheet.title)
        for candidate in candidates:
            if len(candidate) >= 3 and (candidate in normalized_title or normalized_title in candidate):
                score = 0.96
            else:
                score = SequenceMatcher(None, normalized_title, candidate).ratio()
            ranked.append((score, sheet.title, sheet))
    ranked.sort(key=lambda item: (-item[0], item[1].lower()))
    if ranked and ranked[0][0] >= 0.78:
        return ranked[0][2]
    raise ScheduleImportError("sheet_not_found")


def decode_import_csv(file_bytes):
    text = None
    for encoding in ("utf-8-sig", "cp1257", "latin-1"):
        try:
            text = file_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ScheduleImportError("unreadable")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    return list(csv.reader(StringIO(text), dialect))[:MAX_IMPORT_ROWS]


def read_partial_schedule_upload(file_bytes, filename, requested_sheet_name, location, location_id):
    extension = Path(filename or "").suffix.lower()
    if extension not in {".xlsx", ".csv"}:
        raise ScheduleImportError("unsupported")
    if not file_bytes:
        raise ScheduleImportError("empty_file")
    if len(file_bytes) > MAX_IMPORT_BYTES:
        raise ScheduleImportError("too_large")

    if extension == ".csv":
        rows = decode_import_csv(file_bytes)
        sheet_name = "CSV"
    else:
        try:
            workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True, keep_links=False)
            worksheet = select_import_worksheet(workbook, requested_sheet_name, location, location_id)
            sheet_name = worksheet.title
            rows = [list(row[:MAX_IMPORT_COLUMNS]) for _, row in zip(range(MAX_IMPORT_ROWS), worksheet.iter_rows(values_only=True))]
            workbook.close()
        except ScheduleImportError:
            raise
        except Exception as error:
            raise ScheduleImportError("unreadable") from error

    result = parse_partial_schedule_rows(rows, location, location_id)
    result["sheet_name"] = sheet_name
    return result


PREFERRED_WORKER_FILL_COLORS = {
    "worker a": "9FC5E8",
    "worker b": "B6D7A8",
    "worker c": "FFE599",
    "worker d": "D9D2E9",
    "worker e": "F9CB9C",
}
FALLBACK_WORKER_FILL_COLORS = [
    "9FC5E8",
    "B6D7A8",
    "FFE599",
    "D9D2E9",
    "F9CB9C",
    "A2C4C9",
    "D5A6BD",
    "CFE2F3",
]
EMPTY_SHIFT_FILL_COLOR = "B00000"
DATE_FILL_COLOR = "F2F2F2"
WHITE_FILL_COLOR = "FFFFFF"
HEADER_FILL_COLOR = "173D34"
EXPORT_FONT_NAME = "Arial"
THIN_BLACK_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)


def get_export_date_label(schedule_settings, day_number):
    month_name = EXCEL_MONTH_GENITIVE_NAMES.get(schedule_settings["month"], MONTH_NAMES[schedule_settings["month"]].lower())
    return f"{month_name} {day_number}"


def get_export_weekday_letter(schedule_settings, day_number):
    weekday_index = calendar.weekday(schedule_settings["year"], schedule_settings["month"], day_number)
    return EXCEL_WEEKDAY_LETTERS[weekday_index]


def collect_export_worker_names(generated_schedule, worker_summary):
    names = []
    for day in generated_schedule:
        for assignment in day.get("assignments", []):
            worker_name = assignment.get("worker_name")
            if worker_name and worker_name not in names:
                names.append(worker_name)
    for worker in worker_summary:
        worker_name = worker.get("name")
        if worker_name and worker_name not in names:
            names.append(worker_name)
    return names


def build_export_worker_color_map(generated_schedule, worker_summary):
    worker_colors = {}
    fallback_index = 0
    for worker_name in collect_export_worker_names(generated_schedule, worker_summary):
        preferred_color = PREFERRED_WORKER_FILL_COLORS.get(normalize_text(worker_name))
        if preferred_color:
            worker_colors[worker_name] = preferred_color
        else:
            worker_colors[worker_name] = FALLBACK_WORKER_FILL_COLORS[fallback_index % len(FALLBACK_WORKER_FILL_COLORS)]
            fallback_index += 1
    return worker_colors


def get_export_shift_column_count(generated_schedule):
    max_assignments = max((len(day.get("assignments", [])) for day in generated_schedule), default=0)
    return max(2, min(5, max_assignments))


def format_export_assignment(assignment):
    worker_name = assignment.get("worker_name")
    shift_time = assignment.get("shift_time", "")
    return f"{worker_name} {shift_time}".strip() if worker_name else shift_time or "-"


def style_export_cell(cell, fill_color=WHITE_FILL_COLOR, bold=False, font_color="000000"):
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.font = Font(name=EXPORT_FONT_NAME, bold=bold, color=font_color, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    cell.border = THIN_BLACK_BORDER


def create_warning_sheet(workbook, schedule_settings, generated_schedule):
    warning_rows = [day for day in generated_schedule if day.get("warnings")]
    if not warning_rows:
        return

    warning_sheet = workbook.create_sheet("Ispejimai")
    headers = ["Data", "Ispejimai"]
    for column_index, header in enumerate(headers, start=1):
        cell = warning_sheet.cell(row=1, column=column_index, value=header)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
        cell.font = Font(name=EXPORT_FONT_NAME, color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, day in enumerate(warning_rows, start=2):
        date_cell = warning_sheet.cell(row=row_index, column=1, value=get_export_date_label(schedule_settings, day["day"]))
        warning_cell = warning_sheet.cell(row=row_index, column=2, value="\n".join(day["warnings"]))
        date_cell.font = Font(name=EXPORT_FONT_NAME, size=10)
        warning_cell.font = Font(name=EXPORT_FONT_NAME, size=10)
        warning_cell.alignment = Alignment(wrap_text=True, vertical="top")
    autosize_worksheet_columns(warning_sheet)


def create_schedule_workbook(location_name, schedule_settings, generated_schedule, worker_summary):
    workbook = Workbook()
    schedule_sheet = workbook.active
    schedule_sheet.title = "Grafikas"
    schedule_sheet.sheet_view.showGridLines = False

    worker_colors = build_export_worker_color_map(generated_schedule, worker_summary)
    shift_column_count = get_export_shift_column_count(generated_schedule)

    for row_index, day in enumerate(generated_schedule, start=1):
        row_values = [
            day.get("demand_label", ""),
            get_export_weekday_letter(schedule_settings, day["day"]),
            get_export_date_label(schedule_settings, day["day"]),
        ]
        assignments = day.get("assignments", [])[:shift_column_count]
        row_values.extend(format_export_assignment(assignment) for assignment in assignments)
        while len(row_values) < 3 + shift_column_count:
            row_values.append("")

        for column_index, value in enumerate(row_values, start=1):
            cell = schedule_sheet.cell(row=row_index, column=column_index, value=value)
            if column_index <= 2:
                style_export_cell(cell, WHITE_FILL_COLOR, bold=True)
            elif column_index == 3:
                style_export_cell(cell, DATE_FILL_COLOR)
            else:
                assignment_index = column_index - 4
                assignment = assignments[assignment_index] if assignment_index < len(assignments) else None
                if assignment and assignment.get("worker_name"):
                    style_export_cell(cell, worker_colors.get(assignment["worker_name"], FALLBACK_WORKER_FILL_COLORS[0]))
                elif assignment:
                    style_export_cell(cell, EMPTY_SHIFT_FILL_COLOR, font_color="FFFFFF")
                else:
                    style_export_cell(cell, WHITE_FILL_COLOR)
        schedule_sheet.row_dimensions[row_index].height = 18

    schedule_sheet.column_dimensions["A"].width = 4
    schedule_sheet.column_dimensions["B"].width = 4
    schedule_sheet.column_dimensions["C"].width = 15
    for column_index in range(4, 4 + shift_column_count):
        schedule_sheet.column_dimensions[get_column_letter(column_index)].width = 22
    schedule_sheet.freeze_panes = "D1"
    schedule_sheet.page_setup.orientation = "landscape"
    schedule_sheet.page_setup.fitToWidth = 1

    create_warning_sheet(workbook, schedule_settings, generated_schedule)

    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    header_font = Font(name=EXPORT_FONT_NAME, color="FFFFFF", bold=True)

    summary_sheet = workbook.create_sheet("Suvestine")
    summary_headers = ["Vardas", "Etatas", "Pamainu sk.", "Dirbtos valandos", "Reikalingos valandos", "Skirtumas", "Savaitgaliai", "Uzdarymai"]
    for column_index, header in enumerate(summary_headers, start=1):
        cell = summary_sheet.cell(row=1, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
    for row_index, item in enumerate(worker_summary, start=2):
        row_values = [
            item["name"],
            item["etatas"],
            item["assigned_shifts"],
            item["assigned_hours"],
            item["target_hours"],
            item["hours_difference"],
            item["weekend_days"],
            item["closing_shifts"],
        ]
        for column_index, value in enumerate(row_values, start=1):
            summary_sheet.cell(row=row_index, column=column_index, value=value).font = Font(name=EXPORT_FONT_NAME, size=10)
    autosize_worksheet_columns(summary_sheet)

    return workbook


def build_export_filename(location, settings):
    safe_location = re.sub(r"[^a-z0-9-]+", "-", normalize_text(location["name"])).strip("-") or "grafikas"
    return f"{safe_location}-{settings['year']}-{settings['month']:02d}.xlsx"


def save_schedule_export(location):
    workbook = create_schedule_workbook(
        location["name"],
        location["schedule_settings"],
        location["generated_schedule"],
        location["worker_summary"] or [],
    )
    filename = build_export_filename(location, location["schedule_settings"])
    export_path = EXPORTS_DIR / filename
    workbook.save(export_path)
    return export_path


def get_location_rule_overview(location_id):
    rule = get_location_rule(location_id)
    overview = []
    for weekday_index, day_rules in enumerate(rule["days"]):
        full_shifts, half_shifts = split_template_shifts(day_rules)
        overview.append({
            "weekday_name": DAY_LONG_NAMES[weekday_index],
            "full_shifts_text": ", ".join(f"{shift['start']}-{shift['end']} ({shift['hours']:g}h)" for shift in full_shifts) or "-",
            "half_shift_text": ", ".join(f"{shift['start']}-{shift['end']} ({shift['hours']:g}h)" for shift in half_shifts) or "-",
            "default_demand": format_demand_value(len(full_shifts) + (0.5 if half_shifts else 0.0)),
            "breaks_text": ", ".join(day_rules["breaks"]) if day_rules["breaks"] else "-",
        })
    return {"name": rule["name"], "notes": rule["notes"], "days": overview}


def get_location_tabs(active_location_id):
    return [{"id": config["id"], "name": app_data["locations"][config["id"]]["name"], "active": config["id"] == active_location_id} for config in LOCATION_CONFIGS]


def get_location(location_id):
    if location_id not in app_data["locations"]:
        location_id = DEFAULT_LOCATION_ID
    return location_id, app_data["locations"][location_id]


def get_worker_or_404(location, worker_id):
    worker = next((item for item in location["workers"] if item["id"] == worker_id), None)
    if worker is None:
        abort(404)
    return worker


def get_requested_location_id():
    return request.form.get("location_id") or request.args.get("location") or DEFAULT_LOCATION_ID


def clear_generated_results(location):
    location["generated_schedule"] = []
    location["worker_summary"] = None
    location["schedule_insights"] = None


def get_demand_context(location, location_id):
    settings = location["schedule_settings"]
    return parse_daily_demand_map(location["demand_raw"], settings["year"], settings["month"], location_id)


def build_workers_for_view(location, location_id):
    settings = location["schedule_settings"]
    expected_days = get_days_in_month(settings["year"], settings["month"])
    demand_map, _ = get_demand_context(location, location_id)
    workers = []
    for worker in location["workers"]:
        worker_runtime = build_worker_runtime(worker, settings, location_id, demand_map)
        worker_runtime["status"] = get_worker_status(
            worker_runtime["day_count"],
            expected_days,
            worker_runtime["filled_day_count"],
        )
        workers.append(worker_runtime)
    return sort_workers_for_display(workers)


def build_dashboard_stats(location, workers, demand_rows):
    total_requested_hours = sum(row["requested_hours"] for row in demand_rows)
    total_workers = len(workers)
    total_entered_lines = sum(worker["filled_day_count"] for worker in workers)
    expected_total_lines = len(demand_rows) * total_workers
    incomplete_workers = [
        {
            "name": worker["name"],
            "missing_days": max(0, len(demand_rows) - worker["filled_day_count"]),
            "extra_lines": max(0, worker["day_count"] - len(demand_rows)),
        }
        for worker in workers
        if worker["filled_day_count"] != len(demand_rows) or worker["day_count"] > len(demand_rows)
    ]
    stats = {
        "total_workers": total_workers,
        "total_requested_hours": total_requested_hours,
        "entered_lines": total_entered_lines,
        "expected_lines": expected_total_lines,
        "incomplete_workers": incomplete_workers,
        "is_ready": bool(workers) and not incomplete_workers,
        "uncovered_shifts": 0,
        "warning_days": 0,
        "assigned_hours": 0,
    }

    if location["generated_schedule"]:
        uncovered_shifts = 0
        warning_days = 0
        for day in location["generated_schedule"]:
            uncovered_shifts += sum(1 for assignment in day["assignments"] if not assignment["worker_name"])
            if day["warnings"]:
                warning_days += 1
        stats["uncovered_shifts"] = uncovered_shifts
        stats["warning_days"] = warning_days

    if location["worker_summary"]:
        stats["assigned_hours"] = round(sum(item["assigned_hours"] for item in location["worker_summary"]), 1)

    return stats


def build_schedule_progress(generated_schedule):
    assignments = [
        assignment
        for day in generated_schedule if isinstance(day, dict)
        for assignment in day.get("assignments", []) if isinstance(assignment, dict)
    ]
    return {
        "assigned": sum(bool(assignment.get("worker_name")) for assignment in assignments),
        "total": len(assignments),
    }


def apply_schedule_form_assignments(generated_schedule, workers, form_data):
    updated_schedule = deepcopy(generated_schedule)
    workers_by_id = {str(worker.get("id")): worker for worker in workers if worker.get("id")}
    for day_record in updated_schedule:
        if not isinstance(day_record, dict):
            continue
        day = day_record.get("day")
        assignments = day_record.get("assignments", [])
        if not isinstance(assignments, list):
            continue
        for assignment_index, assignment in enumerate(assignments):
            if not isinstance(assignment, dict):
                continue
            time_field_name = f"assignment_time_{day}_{assignment_index}"
            if time_field_name in form_data:
                raw_shift_time = str(form_data.get(time_field_name) or "").strip()
                interval = parse_assignment_interval(raw_shift_time)
                if interval:
                    assignment["shift_time"] = f"{interval[0]}-{interval[1]}"
                    assignment.pop("time_warning", None)
                else:
                    assignment["time_warning"] = (
                        f"Rankinis laikas '{raw_shift_time or '(tuscia)'}' netinkamas; "
                        f"paliktas {assignment.get('shift_time') or 'sablono laikas'}."
                    )
            field_name = f"assignment_{day}_{assignment_index}"
            if field_name not in form_data:
                continue
            worker_id = str(form_data.get(field_name) or "").strip()
            if worker_id == "__preserve__":
                continue
            worker = workers_by_id.get(worker_id)
            if worker:
                assignment["worker_id"] = worker_id
                assignment["worker_name"] = worker["name"]
            else:
                assignment["worker_id"] = None
                assignment["worker_name"] = None
    return updated_schedule


def get_import_notice():
    status = request.args.get("import_status")
    if status == "ok":
        day_count = sanitize_int(request.args.get("imported_days"), 0, 0, 366)
        assigned_count = sanitize_int(request.args.get("imported_assignments"), 0, 0, 5000)
        inferred_count = sanitize_int(request.args.get("inferred_times"), 0, 0, 5000)
        unknown_count = sanitize_int(request.args.get("unknown_workers"), 0, 0, 5000)
        warning_count = sanitize_int(request.args.get("import_warnings"), 0, 0, 5000)
        sheet_name = str(request.args.get("import_sheet") or "").strip()[:80]
        text = f"Importuota {day_count} d. ir {assigned_count} paskirtos pamainos"
        if sheet_name:
            text += f" iš lapo „{sheet_name}“"
        text += "."
        if inferred_count:
            text += f" {inferred_count} tuščių pamainų perėmė artimiausią įkelto grafiko laiką."
        if unknown_count:
            text += f" Neatpažintų darbuotojų: {unknown_count}; jų įrašai palikti peržiūrai."
        if warning_count:
            text += f" Neaiškių pamainos langelių praleista: {warning_count}."
        return {"kind": "success", "text": text}

    if status == "error":
        messages = {
            "missing_file": "Pasirink .xlsx arba .csv grafiko failą.",
            "empty_file": "Įkeltas failas tuščias.",
            "too_large": "Failas per didelis. Didžiausias dydis yra 5 MB.",
            "unsupported": "Tinka tik .xlsx arba .csv failai.",
            "sheet_not_found": "Nepavyko parinkti lokacijos lapo. Įrašyk tikslų lapo pavadinimą ir bandyk dar kartą.",
            "no_rows": "Faile nerasta programoje pasirinkto mėnesio grafiko eilučių.",
            "unreadable": "Failo nepavyko saugiai perskaityti.",
        }
        return {"kind": "error", "text": messages.get(request.args.get("import_error"), messages["unreadable"])}
    return None


def render_home_page(location_id):
    location_id, location = get_location(location_id)
    settings = location["schedule_settings"]
    expected_days = get_days_in_month(settings["year"], settings["month"])
    _, demand_rows = get_demand_context(location, location_id)
    workers = build_workers_for_view(location, location_id)
    return render_template(
        "index.html",
        locations=get_location_tabs(location_id),
        current_location=location,
        current_location_id=location_id,
        workers=workers,
        selected_month=settings["month"],
        selected_year=settings["year"],
        expected_days=expected_days,
        month_name=MONTH_NAMES[settings["month"]],
        location_rule=get_location_rule_overview(location_id),
        demand_rows=demand_rows,
        default_demand_raw=build_default_demand_raw(location_id, settings["year"], settings["month"]),
        dashboard_stats=build_dashboard_stats(location, workers, demand_rows),
        schedule_progress=build_schedule_progress(location["generated_schedule"]),
        import_notice=get_import_notice(),
        build_number=BUILD_NUMBER,
        generated_schedule=location["generated_schedule"],
        worker_summary=location["worker_summary"],
        schedule_insights=location["schedule_insights"],
        ai_status=get_ai_status(),
    )


def render_settings_page(location_id):
    location_id, location = get_location(location_id)
    settings = location["schedule_settings"]
    expected_days = get_days_in_month(settings["year"], settings["month"])
    _, demand_rows = get_demand_context(location, location_id)
    return render_template(
        "settings.html",
        locations=get_location_tabs(location_id),
        current_location=location,
        current_location_id=location_id,
        selected_month=settings["month"],
        selected_year=settings["year"],
        expected_days=expected_days,
        month_name=MONTH_NAMES[settings["month"]],
        demand_rows=demand_rows,
        default_demand_raw=build_default_demand_raw(location_id, settings["year"], settings["month"]),
        build_number=BUILD_NUMBER,
    )


def render_info_page(location_id):
    location_id, location = get_location(location_id)
    return render_template(
        "info.html",
        locations=get_location_tabs(location_id),
        current_location=location,
        current_location_id=location_id,
        build_number=BUILD_NUMBER,
    )


@app.route("/")
def home():
    return render_home_page(get_requested_location_id())


@app.route("/settings")
def settings_page():
    return render_settings_page(get_requested_location_id())


@app.route("/info")
def info_page():
    return render_info_page(get_requested_location_id())


@app.route("/worker/<worker_id>/edit")
def edit_worker_page(worker_id):
    location_id, location = get_location(get_requested_location_id())
    worker = get_worker_or_404(location, worker_id)
    return render_template(
        "edit_worker.html",
        worker=worker,
        locations=get_location_tabs(location_id),
        current_location=location,
        current_location_id=location_id,
        expected_days=get_days_in_month(
            location["schedule_settings"]["year"],
            location["schedule_settings"]["month"],
        ),
        build_number=BUILD_NUMBER,
    )


@app.route("/save_settings", methods=["POST"])
def save_settings():
    location_id, location = get_location(get_requested_location_id())
    month = request.form.get("month", "").strip()
    year = request.form.get("year", "").strip()
    full_time_hours = request.form.get("full_time_hours", "").strip()
    current_settings = location["schedule_settings"]
    location["schedule_settings"]["month"] = sanitize_int(month, current_settings["month"], 1, 12)
    location["schedule_settings"]["year"] = sanitize_int(year, current_settings["year"], 2000, 2100)
    location["schedule_settings"]["full_time_hours"] = sanitize_int(
        full_time_hours,
        current_settings["full_time_hours"],
        1,
        250,
    )
    location["demand_raw"] = sanitize_demand_raw(request.form.get("demand_raw", ""), location_id, location["schedule_settings"])
    clear_generated_results(location)
    save_app_data()
    return redirect(url_for("home", location=location_id))


@app.route("/add_worker", methods=["POST"])
def add_worker():
    location_id, location = get_location(get_requested_location_id())
    name = request.form.get("name", "").strip()
    etatas = request.form.get("etatas", "").strip()
    availability_raw = str(request.form.get("availability", "") or "")
    if name and etatas:
        location["workers"].append({"id": uuid4().hex, "name": name, "etatas": etatas, "availability_raw": availability_raw})
        clear_generated_results(location)
        save_app_data()
    return redirect(url_for("home", location=location_id))


@app.route("/delete_worker/<worker_id>", methods=["POST"])
def delete_worker(worker_id):
    location_id, location = get_location(get_requested_location_id())
    original_count = len(location["workers"])
    location["workers"] = [worker for worker in location["workers"] if worker["id"] != worker_id]
    if len(location["workers"]) != original_count:
        clear_generated_results(location)
        save_app_data()
    return redirect(url_for("home", location=location_id))


@app.route("/worker/<worker_id>/update", methods=["POST"])
def update_worker(worker_id):
    location_id, location = get_location(get_requested_location_id())
    worker = get_worker_or_404(location, worker_id)
    name = request.form.get("name", "").strip()
    etatas = request.form.get("etatas", "").strip()
    availability_raw = str(request.form.get("availability", "") or "")
    if name and etatas:
        worker.update({"name": name, "etatas": etatas, "availability_raw": availability_raw})
        clear_generated_results(location)
        save_app_data()
    return redirect(url_for("home", location=location_id, _anchor="workers"))


@app.route("/generate_schedule", methods=["POST"])
def generate_schedule_route():
    location_id, location = get_location(get_requested_location_id())
    demand_map, _ = get_demand_context(location, location_id)
    workers = build_workers_for_view(location, location_id)
    fill_open_slots = request.form.get("mode") != "draft"
    generated_schedule, worker_summary = generate_month_schedule(
        workers,
        location["schedule_settings"],
        location_id,
        demand_map,
        fill_open_slots=fill_open_slots,
    )
    location["generated_schedule"] = generated_schedule
    location["worker_summary"] = worker_summary
    location["schedule_insights"] = build_schedule_insights(location, generated_schedule, worker_summary)
    save_app_data()
    return redirect(url_for("home", location=location_id, _anchor="schedule"))


@app.route("/import_partial_schedule", methods=["POST"])
def import_partial_schedule_route():
    location_id, location = get_location(get_requested_location_id())
    upload = request.files.get("schedule_file")
    if not upload or not upload.filename:
        return redirect(url_for("home", location=location_id, import_status="error", import_error="missing_file", _anchor="schedule-import"))

    try:
        file_bytes = upload.read(MAX_IMPORT_BYTES + 1)
        import_result = read_partial_schedule_upload(
            file_bytes,
            upload.filename,
            request.form.get("sheet_name", "").strip(),
            location,
            location_id,
        )
    except ScheduleImportError as error:
        return redirect(url_for("home", location=location_id, import_status="error", import_error=error.code, _anchor="schedule-import"))
    except OSError:
        return redirect(url_for("home", location=location_id, import_status="error", import_error="unreadable", _anchor="schedule-import"))

    demand_map, _ = get_demand_context(location, location_id)
    demand_map.update(import_result["demand_updates"])
    location["demand_raw"] = "\n".join(
        format_demand_value(demand_map[day])
        for day in range(1, get_days_in_month(location["schedule_settings"]["year"], location["schedule_settings"]["month"]) + 1)
    )
    workers = build_workers_for_view(location, location_id)
    generated_schedule, worker_summary = generate_month_schedule(
        workers,
        location["schedule_settings"],
        location_id,
        demand_map,
        existing_schedule=import_result["schedule"],
        fill_open_slots=False,
    )
    location["generated_schedule"] = generated_schedule
    location["worker_summary"] = worker_summary
    location["schedule_insights"] = build_schedule_insights(location, generated_schedule, worker_summary)
    save_app_data()
    return redirect(url_for(
        "home",
        location=location_id,
        import_status="ok",
        imported_days=import_result["day_count"],
        imported_assignments=import_result["assigned_count"],
        inferred_times=import_result["inferred_time_count"],
        unknown_workers=import_result["unknown_worker_count"],
        import_warnings=import_result["warning_count"],
        import_sheet=import_result["sheet_name"],
        _anchor="schedule-import",
    ))


def update_existing_schedule(location_id, location, fill_open_slots):
    if not location["generated_schedule"]:
        abort(400)
    demand_map, _ = get_demand_context(location, location_id)
    workers = build_workers_for_view(location, location_id)
    edited_schedule = apply_schedule_form_assignments(location["generated_schedule"], workers, request.form)
    generated_schedule, worker_summary = generate_month_schedule(
        workers,
        location["schedule_settings"],
        location_id,
        demand_map,
        existing_schedule=edited_schedule,
        fill_open_slots=fill_open_slots,
    )
    location["generated_schedule"] = generated_schedule
    location["worker_summary"] = worker_summary
    location["schedule_insights"] = build_schedule_insights(location, generated_schedule, worker_summary)
    save_app_data()


@app.route("/save_partial_schedule", methods=["POST"])
def save_partial_schedule_route():
    location_id, location = get_location(get_requested_location_id())
    update_existing_schedule(location_id, location, fill_open_slots=False)
    return redirect(url_for("home", location=location_id, _anchor="schedule"))


@app.route("/complete_schedule", methods=["POST"])
def complete_schedule_route():
    location_id, location = get_location(get_requested_location_id())
    update_existing_schedule(location_id, location, fill_open_slots=True)
    return redirect(url_for("home", location=location_id, _anchor="schedule"))


@app.route("/export_schedule")
def export_schedule():
    location_id, location = get_location(get_requested_location_id())
    if not location["generated_schedule"]:
        abort(404)

    export_path = save_schedule_export(location)
    return send_file(
        export_path,
        as_attachment=True,
        download_name=export_path.name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(
        host=os.getenv("HOST", "127.0.0.1"),
        port=sanitize_int(os.getenv("PORT"), 5000, 1, 65535),
        debug=os.getenv("FLASK_DEBUG") == "1",
    )
