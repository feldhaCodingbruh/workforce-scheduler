import unittest
from copy import deepcopy
from datetime import datetime
from io import BytesIO
from unittest.mock import patch

from openpyxl import Workbook

import app


class AvailabilityParserTests(unittest.TestCase):
    def test_normalize_text_removes_lithuanian_diacritics(self):
        self.assertEqual(app.normalize_text("Rytinė norėčiau"), "rytine noreciau")
        self.assertEqual(app.normalize_text("Galiu šeštadienį"), "galiu sestadieni")

    def test_morning_preference_with_diacritics_is_recognized(self):
        parsed = app.parse_single_line_rule_based("Rytinė norėčiau")

        self.assertEqual(parsed["type"], "available")
        self.assertEqual(parsed["preference"], "morning")

    def test_common_lithuanian_availability_formats(self):
        cases = {
            "galiu nuo 16:40": ("from_time", "16:40", None),
            "iki 17": ("until_time", None, "17:00"),
            "10-18": ("time_range", "10:00", "18:00"),
            "negaliu (egzas)": ("unavailable", None, None),
        }

        for text, expected in cases.items():
            with self.subTest(text=text):
                parsed = app.parse_single_line_rule_based(text)
                self.assertEqual(
                    (parsed["type"], parsed["from_time"], parsed["until_time"]),
                    expected,
                )

    def test_split_availability_is_recognized(self):
        parsed = app.parse_single_line_rule_based("iki14/nuo19")

        self.assertEqual(parsed["type"], "split")
        self.assertEqual(parsed["until_time"], "14:00")
        self.assertEqual(parsed["second_from_time"], "19:00")

    def test_blank_lines_preserve_calendar_day_positions(self):
        lines = app.parse_availability_lines("galiu\n\nnegaliu\nnuo 17\n")

        self.assertEqual(lines, ["galiu", "", "negaliu", "nuo 17"])

    def test_blank_day_does_not_count_as_complete_input(self):
        self.assertEqual(app.get_worker_status(3, 3, 2), "Truksta 1 d.")
        self.assertEqual(app.get_worker_status(3, 3, 3), "Gerai")


class ScheduleRegressionTests(unittest.TestCase):
    def test_demand_above_template_creates_override_full_shift(self):
        shifts, warnings = app.build_requested_shifts("location-g", 2026, 6, 1, 3.0)

        self.assertEqual(len(shifts), 3)
        self.assertEqual([shift["slot_kind"] for shift in shifts], ["Pilna", "Pilna", "Pilna+"])
        self.assertEqual(shifts[2]["start"], "13:00")
        self.assertEqual(shifts[2]["end"], "21:30")
        self.assertTrue(shifts[2]["template_override"])
        self.assertIn("Papildomos pamainos kuriamos kaip override", warnings[0])

    def test_schedule_generation_assigns_override_full_shift(self):
        settings = {"year": 2026, "month": 6, "full_time_hours": 160}
        demand = {day: (3.0 if day == 1 else 0.0) for day in range(1, 31)}
        workers = [
            {"id": "a", "name": "Worker A", "etatas": "1.0", "availability_raw": "\n".join(["galiu"] * 30)},
            {"id": "b", "name": "Worker B", "etatas": "1.0", "availability_raw": "\n".join(["galiu"] * 30)},
            {"id": "c", "name": "Worker C", "etatas": "1.0", "availability_raw": "\n".join(["galiu"] * 30)},
        ]
        runtime_workers = [
            app.build_worker_runtime(worker, settings, "location-g", demand)
            for worker in workers
        ]

        schedule, _ = app.generate_month_schedule(runtime_workers, settings, "location-g", demand)
        day_one = schedule[0]

        self.assertEqual(len(day_one["assignments"]), 3)
        self.assertEqual([item["slot_kind"] for item in day_one["assignments"]], ["Pilna", "Pilna", "Pilna+"])
        self.assertTrue(all(item["worker_name"] for item in day_one["assignments"]))
        self.assertFalse(any("Nera darbuotojo" in warning for warning in day_one["warnings"]))

    def test_monthly_target_uses_configured_full_time_hours(self):
        self.assertEqual(app.etatas_to_month_hours("0.5", 168), 84)
        self.assertEqual(app.etatas_to_month_hours("0.75", 168), 126)

    def test_turnaround_keeps_11h_hard_floor_and_prefers_13h_buffer(self):
        prior_assignment = {1: {"start": 13 * 60, "end": 21 * 60 + 30}}

        self.assertTrue(
            app.would_break_rest_gap(prior_assignment, 2, "08:00", "16:30")
        )
        self.assertFalse(
            app.would_break_rest_gap(prior_assignment, 2, "08:30", "17:00")
        )
        self.assertAlmostEqual(
            app.get_soft_turnaround_penalty(
                prior_assignment,
                2,
                "10:00",
                "18:30",
            ),
            0.5,
        )
        self.assertEqual(
            app.get_soft_turnaround_penalty(
                prior_assignment,
                2,
                "13:00",
                "21:30",
            ),
            0,
        )

    def test_adjacent_start_bonus_is_small_and_time_bounded(self):
        prior_assignment = {1: {"start": 10 * 60, "end": 18 * 60 + 30}}

        self.assertEqual(
            app.get_adjacent_start_bonus(prior_assignment, 2, "11:00"),
            app.ADJACENT_START_BONUS,
        )
        self.assertEqual(
            app.get_adjacent_start_bonus(prior_assignment, 2, "11:30"),
            0,
        )

    def test_generation_prefers_rested_opener_without_blocking_coverage(self):
        settings = {"year": 2026, "month": 6, "full_time_hours": 160}
        demand = {
            day: (2.0 if day == 1 else 1.0 if day == 2 else 0.0)
            for day in range(1, 31)
        }
        workers = [
            {
                "id": "a",
                "name": "Worker A",
                "etatas": "1.0",
                "availability_raw": "nuo 13\ngaliu",
            },
            {
                "id": "b",
                "name": "Worker B",
                "etatas": "1.0",
                "availability_raw": "galiu\ngaliu",
            },
        ]
        runtime_workers = [
            app.build_worker_runtime(worker, settings, "location-g", demand)
            for worker in workers
        ]

        schedule, _ = app.generate_month_schedule(
            runtime_workers,
            settings,
            "location-g",
            demand,
        )

        self.assertEqual(
            [item["worker_name"] for item in schedule[0]["assignments"]],
            ["Worker B", "Worker A"],
        )
        self.assertEqual(schedule[1]["assignments"][0]["worker_name"], "Worker B")

    def test_worker_with_partial_availability_stays_in_summary(self):
        settings = {"year": 2026, "month": 6, "full_time_hours": 160}
        demand = {day: (1.0 if day == 1 else 0.0) for day in range(1, 31)}
        workers = [
            {"id": "a", "name": "Worker A", "etatas": "0.5", "availability_raw": "galiu"},
            {"id": "b", "name": "Worker B", "etatas": "0.5", "availability_raw": "galiu"},
        ]
        runtime_workers = [
            app.build_worker_runtime(worker, settings, "location-d", demand)
            for worker in workers
        ]

        _, summary = app.generate_month_schedule(runtime_workers, settings, "location-d", demand)

        self.assertEqual({item["name"] for item in summary}, {"Worker A", "Worker B"})

    def test_partial_schedule_completion_preserves_existing_assignment(self):
        settings = {"year": 2026, "month": 6, "full_time_hours": 160}
        demand = {day: (2.0 if day == 1 else 0.0) for day in range(1, 31)}
        workers = [
            {"id": "a", "name": "Worker A", "etatas": "1.0", "availability_raw": "galiu"},
            {"id": "b", "name": "Worker B", "etatas": "1.0", "availability_raw": "galiu"},
        ]
        runtime_workers = [
            app.build_worker_runtime(worker, settings, "location-g", demand)
            for worker in workers
        ]
        partial_schedule, _ = app.generate_month_schedule(
            runtime_workers,
            settings,
            "location-g",
            demand,
            fill_open_slots=False,
        )
        partial_schedule[0]["assignments"][0].update(
            {"worker_id": "a", "worker_name": "Worker A"}
        )

        completed_schedule, summary = app.generate_month_schedule(
            runtime_workers,
            settings,
            "location-g",
            demand,
            existing_schedule=partial_schedule,
        )

        day_one = completed_schedule[0]
        self.assertEqual(day_one["assignments"][0]["worker_name"], "Worker A")
        self.assertEqual(day_one["assignments"][0]["worker_id"], "a")
        self.assertEqual(day_one["assignments"][1]["worker_name"], "Worker B")
        self.assertFalse(any(not item["worker_name"] for item in day_one["assignments"]))
        self.assertEqual(sum(item["assigned_shifts"] for item in summary), 2)

    def test_manual_assignment_is_preserved_when_availability_disagrees(self):
        settings = {"year": 2026, "month": 6, "full_time_hours": 160}
        demand = {day: (1.0 if day == 1 else 0.0) for day in range(1, 31)}
        workers = [
            {"id": "a", "name": "Worker A", "etatas": "1.0", "availability_raw": "negaliu"},
            {"id": "b", "name": "Worker B", "etatas": "1.0", "availability_raw": "galiu"},
        ]
        runtime_workers = [
            app.build_worker_runtime(worker, settings, "location-g", demand)
            for worker in workers
        ]
        partial_schedule = [{
            "day": 1,
            "assignments": [{
                "shift_label": "Pilna 1",
                "slot_kind": "Pilna",
                "shift_time": "10:00-18:30",
                "worker_id": "a",
                "worker_name": "Worker A",
            }],
        }]

        completed_schedule, _ = app.generate_month_schedule(
            runtime_workers,
            settings,
            "location-g",
            demand,
            existing_schedule=partial_schedule,
        )

        self.assertEqual(completed_schedule[0]["assignments"][0]["worker_name"], "Worker A")
        self.assertTrue(
            any("Rankinis pasirinkimas neatitinka" in warning for warning in completed_schedule[0]["warnings"])
        )

    def test_schedule_form_uses_worker_ids_and_can_clear_a_slot(self):
        schedule = [{
            "day": 1,
            "assignments": [
                {"worker_id": None, "worker_name": None},
                {"worker_id": "a", "worker_name": "Worker A"},
            ],
        }]
        workers = [
            {"id": "a", "name": "Worker A"},
            {"id": "b", "name": "Worker B"},
        ]

        updated = app.apply_schedule_form_assignments(
            schedule,
            workers,
            {"assignment_1_0": "b", "assignment_1_1": ""},
        )

        self.assertEqual(updated[0]["assignments"][0]["worker_name"], "Worker B")
        self.assertEqual(updated[0]["assignments"][0]["worker_id"], "b")
        self.assertIsNone(updated[0]["assignments"][1]["worker_name"])
        self.assertEqual(schedule[0]["assignments"][0]["worker_name"], None)


class ManualTimeAndImportTests(unittest.TestCase):
    def test_schedule_form_accepts_flexible_custom_time(self):
        schedule = [{
            "day": 1,
            "assignments": [{
                "shift_time": "13:00-21:30",
                "worker_id": None,
                "worker_name": None,
            }],
        }]

        updated = app.apply_schedule_form_assignments(
            schedule,
            [],
            {"assignment_time_1_0": "13.30 \u2013 21.30"},
        )

        self.assertEqual(updated[0]["assignments"][0]["shift_time"], "13:30-21:30")

    def test_import_keeps_exact_times_and_infers_them_for_later_blank_day(self):
        location_id = app.DEFAULT_LOCATION_ID
        location_name = app.LOCATION_CONFIGS[0]["name"]
        settings = {"year": 2026, "month": 6, "full_time_hours": 160}
        workers = [
            {"id": "a", "name": "Worker A", "etatas": "1.0", "availability_raw": "\n".join(["galiu"] * 30)},
            {"id": "b", "name": "Worker B", "etatas": "1.0", "availability_raw": "\n".join(["galiu"] * 30)},
            {"id": "c", "name": "Worker C", "etatas": "1.0", "availability_raw": "\n".join(["galiu"] * 30)},
        ]
        location = {
            "name": location_name,
            "schedule_settings": settings,
            "demand_raw": "\n".join(["0"] * 30),
            "workers": workers,
        }
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = location_name
        sheet.append([
            2.5,
            "P",
            datetime(2026, 6, 1),
            "Worker A 10:00-18:30",
            "Worker B 13:30-21:30",
            "Worker C 14:00-21:30",
        ])
        sheet.append([2.5, "P", datetime(2026, 6, 8), None, None, None])
        file_buffer = BytesIO()
        workbook.save(file_buffer)

        imported = app.read_partial_schedule_upload(
            file_buffer.getvalue(),
            "partial.xlsx",
            "",
            location,
            location_id,
        )

        self.assertEqual(imported["sheet_name"], location_name)
        self.assertEqual(imported["assigned_count"], 3)
        self.assertEqual(imported["inferred_time_count"], 3)
        self.assertEqual(
            [item["shift_time"] for item in imported["schedule"][1]["assignments"]],
            ["10:00-18:30", "13:30-21:30", "14:00-21:30"],
        )

        demand = {day: (2.5 if day in {1, 8} else 0.0) for day in range(1, 31)}
        runtime_workers = [
            app.build_worker_runtime(worker, settings, location_id, demand)
            for worker in workers
        ]
        completed, _ = app.generate_month_schedule(
            runtime_workers,
            settings,
            location_id,
            demand,
            existing_schedule=imported["schedule"],
            fill_open_slots=True,
        )

        self.assertEqual(
            [item["shift_time"] for item in completed[7]["assignments"]],
            ["10:00-18:30", "13:30-21:30", "14:00-21:30"],
        )
        self.assertTrue(all(item["worker_name"] for item in completed[7]["assignments"]))


class WorkerEditingTests(unittest.TestCase):
    def setUp(self):
        self.location = app.app_data["locations"]["location-a"]
        self.original_location = deepcopy(self.location)

    def tearDown(self):
        self.location.clear()
        self.location.update(self.original_location)

    def test_worker_can_be_updated_without_writing_real_data(self):
        self.location["workers"] = [
            {"id": "worker-1", "name": "Old", "etatas": "0.5", "availability_raw": "galiu"}
        ]
        self.location["generated_schedule"] = [{"day": 1}]
        client = app.app.test_client()

        with patch.object(app, "save_app_data"):
            response = client.post(
                "/worker/worker-1/update",
                data={
                    "location_id": "location-a",
                    "name": "New",
                    "etatas": "0.75",
                    "availability": "galiu\n\nnegaliu",
                },
            )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.location["workers"][0]["name"], "New")
        self.assertEqual(self.location["workers"][0]["etatas"], "0.75")
        self.assertEqual(self.location["workers"][0]["availability_raw"], "galiu\n\nnegaliu")
        self.assertEqual(self.location["generated_schedule"], [])


class PartialScheduleRouteTests(unittest.TestCase):
    def setUp(self):
        self.location = app.app_data["locations"]["location-a"]
        self.original_location = deepcopy(self.location)
        settings = {"year": 2026, "month": 6, "full_time_hours": 160}
        demand_values = ["2"] + ["0"] * 29
        workers = [
            {"id": "a", "name": "Worker A", "etatas": "1.0", "availability_raw": "galiu"},
            {"id": "b", "name": "Worker B", "etatas": "1.0", "availability_raw": "galiu"},
        ]
        self.location.update({
            "schedule_settings": settings,
            "demand_raw": "\n".join(demand_values),
            "workers": workers,
            "worker_summary": None,
            "schedule_insights": None,
        })
        demand, _ = app.get_demand_context(self.location, "location-a")
        runtime_workers = [
            app.build_worker_runtime(worker, settings, "location-a", demand)
            for worker in workers
        ]
        self.location["generated_schedule"], _ = app.generate_month_schedule(
            runtime_workers,
            settings,
            "location-a",
            demand,
            fill_open_slots=False,
        )

    def tearDown(self):
        self.location.clear()
        self.location.update(self.original_location)

    def test_complete_route_saves_form_selection_and_fills_only_blanks(self):
        client = app.app.test_client()

        with patch.object(app, "save_app_data"), patch.object(
            app,
            "build_schedule_insights",
            return_value={"items": [], "ai_used": False, "ai_status": ""},
        ):
            response = client.post(
                "/complete_schedule",
                data={"location_id": "location-a", "assignment_1_0": "a"},
            )

        self.assertEqual(response.status_code, 302)
        day_one = self.location["generated_schedule"][0]
        self.assertEqual(day_one["assignments"][0]["worker_name"], "Worker A")
        self.assertEqual(day_one["assignments"][1]["worker_name"], "Worker B")
        self.assertEqual(response.headers["Location"], "/?location=location-a#schedule")

    def test_import_route_loads_partial_workbook_without_completing_it(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self.location["name"]
        sheet.append([
            2,
            "P",
            datetime(2026, 6, 1),
            "Worker A 10:00-18:30",
            "13:30-21:30",
        ])
        file_buffer = BytesIO()
        workbook.save(file_buffer)
        client = app.app.test_client()

        with patch.object(app, "save_app_data"), patch.object(
            app,
            "build_schedule_insights",
            return_value={"items": [], "ai_used": False, "ai_status": ""},
        ):
            response = client.post(
                "/import_partial_schedule",
                data={
                    "location_id": "location-a",
                    "schedule_file": (BytesIO(file_buffer.getvalue()), "partial.xlsx"),
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(response.status_code, 302)
        day_one = self.location["generated_schedule"][0]
        self.assertEqual(day_one["assignments"][0]["worker_name"], "Worker A")
        self.assertEqual(day_one["assignments"][1]["worker_name"], None)
        self.assertEqual(day_one["assignments"][1]["shift_time"], "13:30-21:30")
        self.assertIn("import_status=ok", response.headers["Location"])

    def test_partial_schedule_editor_renders_worker_selects_and_actions(self):
        client = app.app.test_client()

        response = client.get("/?location=location-a")
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn('name="assignment_1_0"', html)
        self.assertIn('name="assignment_time_1_0"', html)
        self.assertIn('name="schedule_file"', html)
        self.assertIn("Išsaugoti dalį", html)
        self.assertIn("Užbaigti grafiką", html)


class WorkbookExportTests(unittest.TestCase):
    def test_compact_schedule_export_matches_manager_view(self):
        settings = {"year": 2026, "month": 6, "full_time_hours": 160}
        generated_schedule = [
            {
                "day": 12,
                "weekday_name": "Pn",
                "demand_label": "5",
                "assignments": [
                    {"slot_kind": "Pilna", "shift_time": "10:00-17:00", "worker_name": "Worker A"},
                    {"slot_kind": "Pilna", "shift_time": "13:00-21:30", "worker_name": "Worker B"},
                ],
                "warnings": [],
            },
            {
                "day": 13,
                "weekday_name": "St",
                "demand_label": "3",
                "assignments": [
                    {"slot_kind": "Pilna", "shift_time": "10:00-16:00", "worker_name": "Worker B"},
                    {"slot_kind": "Pilna", "shift_time": "13:30-21:30", "worker_name": None},
                ],
                "warnings": ["Nera darbuotojo Pilna 2"],
            },
        ]

        workbook = app.create_schedule_workbook(
            "Location A",
            settings,
            generated_schedule,
            [
                {
                    "name": "Worker A",
                    "etatas": "0.75",
                    "assigned_shifts": 1,
                    "assigned_hours": 7,
                    "target_hours": 120,
                    "hours_difference": -113,
                    "weekend_days": 0,
                    "closing_shifts": 0,
                }
            ],
        )
        sheet = workbook["Grafikas"]

        self.assertEqual(sheet["A1"].value, "5")
        self.assertEqual(sheet["B1"].value, "P")
        self.assertEqual(sheet["C1"].value, "birželio 12")
        self.assertEqual(sheet["D1"].value, "Worker A 10:00-17:00")
        self.assertEqual(sheet["E1"].value, "Worker B 13:00-21:30")
        self.assertEqual(sheet["D1"].fill.fgColor.rgb, "009FC5E8")
        self.assertEqual(sheet["E2"].value, "13:30-21:30")
        self.assertEqual(sheet["E2"].fill.fgColor.rgb, "00B00000")
        self.assertIn("Ispejimai", workbook.sheetnames)
        self.assertEqual(sheet["D1"].font.name, "Arial")
        self.assertEqual(workbook["Ispejimai"]["B2"].font.name, "Arial")
        self.assertEqual(workbook["Suvestine"]["A2"].font.name, "Arial")


if __name__ == "__main__":
    unittest.main()
